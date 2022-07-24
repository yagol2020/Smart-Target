import os
import json
from enum import Enum
import re
import shutil
from typing import Dict, List
from openpyxl import load_workbook
import pandas
import matplotlib.pyplot as plt
import seaborn
import loguru
from crytic_compile import CryticCompile
from smart_target_config import SKIP_SOL_CONTRACT_NAME_PAIR
import evm_cfg_builder.cfg as cfg_builder

mpl_config = {
    "font.family": "sans-serif",  # 使用衬线体
    "font.sans-serif": ["SimHei"],  # 全局默认使用衬线宋体
    "font.size": 14,  # 五号，10.5磅
    "axes.unicode_minus": False,
    "mathtext.fontset": "cm",  # 设置 LaTeX 字体，stix 近似于 Times 字体
}
# plt.rcParams['font.sans-serif']=['SimHei','DejaVu Sans'] #用来正常显示中文标签
plt.rcParams.update(mpl_config)


class Summary():
    """
    执行相关信息
    """

    def __init__(self, slither_time, target_gen_time, mythril_target_time, mythril_fully_time, chose_method, success,
                 reason, target_time_out, fully_time_out) -> None:
        self.slither_time = slither_time  # slither静态检测的时间
        self.target_gen_time = target_gen_time  # 生成制导信息的时间
        self.mythril_target_time = mythril_target_time  # mythril制导的时间
        self.mythril_fully_time = mythril_fully_time  # mythril全量的时间
        # 字节码映射源代码选择的方法.1为src map.2为ASM
        self.chose_method = chose_method
        self.success = success  # 是否成功.目前否的情况是制导信息没有生成.比如因为cfg的错误
        self.reason = reason  # 不成功的原因
        self.target_time_out = target_time_out  # 制导符号执行是否超时
        self.fully_time_out = fully_time_out  # 全量符号执行是否超时
        assert self.chose_method == 1, "使用的是2号策略.ASM策略"  # 过滤掉ASM策略.这个策略不够严谨
        assert self.success == True, "目标制导信息生成失败"
        assert self.target_time_out == True, "目标制导符号执行超时"
        assert self.fully_time_out == True, "全量执行超时"


class Mode(Enum):
    """
    符号执行模式
    """
    TARGET = 1
    FULLY = 2


class BugInfoDetectMode(Enum):
    ONE_TIME = 1  # 所有漏洞同时设为目标
    EACH_TIME = 2  # 将每个漏洞单独设为目标


class Tool(Enum):
    """
    工具
    """
    SLITHER = 1
    MYTHRIL_TARGET = 2
    MYTHRIL_FULLY = 3


class BugType(Enum):
    """
    基础漏洞分类.与smartbugs数据集一一对应

    """
    ACCESS_CONTROL = 1
    ARITHMETIC = 2
    BAD_RANDOMNESS = 3
    DENIAL_OF_SERVICE = 4
    FRONT_RUNNING = 5
    OTHER = 6
    REENTRANCY = 7
    SHORT_ADDRESSES = 8
    TIME_MANIPULATION = 9
    UNCHECKED_LOW_LEVEL_CALLS = 10
    I_DO_NOT_KNOW = 11
    SKIP = 12


def convert(origin):
    """
    漏洞映射关系.将slither和mythil的漏洞映射到基础漏洞分类里

    """
    if origin == 'Integer Arithmetic Bugs':
        return BugType.ARITHMETIC
    if origin == "Dependence on predictable environment variable":
        return BugType.TIME_MANIPULATION
    if origin == "External Call To User-Supplied Address":
        return BugType.REENTRANCY
    if origin == "other":
        return BugType.OTHER
    if origin == "time_manipulation":
        return BugType.TIME_MANIPULATION
    if origin == "front_running":
        return BugType.FRONT_RUNNING
    if origin == "Unprotected Ether Withdrawal":  # mythril
        return BugType.I_DO_NOT_KNOW
    if origin == "Multiple Calls in a Single Transaction":
        return BugType.I_DO_NOT_KNOW
    if origin == "reentrancy":
        return BugType.REENTRANCY
    if origin == "Unchecked return value from external call.":
        return BugType.UNCHECKED_LOW_LEVEL_CALLS
    if origin == "Dependence on tx.origin":  # mythril
        return BugType.ACCESS_CONTROL
    if origin == "Exception State":
        return BugType.I_DO_NOT_KNOW
    if origin == "controlled-array-length":
        return BugType.I_DO_NOT_KNOW
    if origin == "weak-prng":
        return BugType.BAD_RANDOMNESS
    if origin == "incorrect-equality":
        return BugType.I_DO_NOT_KNOW
    if origin == "timestamp":
        return BugType.TIME_MANIPULATION
    if origin == "uninitialized-storage":
        return BugType.I_DO_NOT_KNOW
    if origin == "arbitrary-send":
        return BugType.ACCESS_CONTROL
    if origin == "unchecked-send":
        return BugType.UNCHECKED_LOW_LEVEL_CALLS
    if origin == "tautology":  # slither, 权限控制?
        return BugType.I_DO_NOT_KNOW
    if origin == "events-maths":  # slither
        return BugType.I_DO_NOT_KNOW
    if origin == "reentrancy-eth":
        return BugType.REENTRANCY
    if origin == "missing-zero-check":  # slither
        return BugType.I_DO_NOT_KNOW
    if origin == "unchecked-lowlevel":
        return BugType.UNCHECKED_LOW_LEVEL_CALLS
    if origin == "incorrect-modifier":
        return BugType.I_DO_NOT_KNOW
    if origin == "shadowing-state":
        return BugType.I_DO_NOT_KNOW
    if origin == "unchecked_low_level_calls":
        return BugType.UNCHECKED_LOW_LEVEL_CALLS
    if origin == "reentrancy-benign":  # slither的重入类型之一, 但是smartbugs认为没有这样的, 若启动, 则全部为fp
        return BugType.I_DO_NOT_KNOW
    if origin == "reentrancy-events":
        return BugType.I_DO_NOT_KNOW
    if origin == "reentrancy-no-eth":
        return BugType.REENTRANCY
    if origin == "constant-function-asm":
        return BugType.I_DO_NOT_KNOW
    if origin == "events-access":
        return BugType.I_DO_NOT_KNOW
    if origin == "unchecked-transfer":
        return BugType.UNCHECKED_LOW_LEVEL_CALLS
    if origin == "unused-return":  # slither
        return BugType.I_DO_NOT_KNOW
    if origin == "locked-ether":
        return BugType.I_DO_NOT_KNOW
    if origin == "shadowing-local":
        return BugType.I_DO_NOT_KNOW
    if origin == "calls-loop":
        return BugType.DENIAL_OF_SERVICE
    if origin == "uninitialized-state":
        return BugType.I_DO_NOT_KNOW
    if origin == "shadowing-builtin":
        return BugType.I_DO_NOT_KNOW
    if origin == "divide-before-multiply":
        return BugType.I_DO_NOT_KNOW
    if origin == "access_control":
        return BugType.ACCESS_CONTROL
    if origin == "suicidal":
        return BugType.ACCESS_CONTROL
    if origin == "uninitialized-local":
        return BugType.I_DO_NOT_KNOW
    if origin == "Delegatecall to user-supplied address":
        return BugType.ACCESS_CONTROL
    if origin == "controlled-delegatecall":
        return BugType.ACCESS_CONTROL
    if origin == "Unprotected Selfdestruct":  # mythril
        return BugType.ACCESS_CONTROL
    if origin == "Write to an arbitrary storage location":  # mythril的类型, 是权限控制么?
        return BugType.I_DO_NOT_KNOW
    if origin == "tx-origin":
        return BugType.ACCESS_CONTROL
    if origin == "erc20-interface":
        return BugType.I_DO_NOT_KNOW
    if origin == "denial_of_service":
        return BugType.DENIAL_OF_SERVICE
    if origin == "arithmetic":
        return BugType.ARITHMETIC
    if origin == "bad_randomness":
        return BugType.BAD_RANDOMNESS
    if origin == "short_addresses":
        return BugType.SHORT_ADDRESSES
    raise Exception(origin + " NOT FIND")


class BugInfoState(Enum):
    """
    单个漏洞报告的结果状态.是否误报、漏报
    """
    TP = 1  # 正确
    FP = 2  # 误报
    FN = 3  # 漏报
    UN = 4  # 未知


class BugInfo:
    """
    漏洞信息
    """

    def __init__(self, line_num, bug_type: BugType, origin):
        self.line_num = line_num  # 漏洞所在行号
        self.bug_type = bug_type  # 漏洞基础类型
        self.origin = origin  # 漏洞原始名称
        if self.bug_type == BugType.I_DO_NOT_KNOW:
            self.state = BugInfoState.UN  # 该漏洞状态.因为不能映射到基础类型.所以UN
        else:
            self.state = BugInfoState.FP  # 能映射到基础类型.假定全部为误报.在各个报告中和truth比较能够修改为其他状态

    def __str__(self):
        """
        友好的输出
        """
        return str(self.line_num) + ": " + str(self.bug_type) + ": \"" + str(self.origin) + "\""

    def __repr__(self):
        """
        友好的输出
        """
        return str(self.line_num) + ": " + str(self.bug_type) + ": \"" + str(self.origin) + "\""

    def to_json_str(self):
        """
        转换为json字符串.用于输出报告
        """
        return json.dumps({
            "line_num": self.line_num,
            "bug_type": self.bug_type.name,
            "origin": self.origin,
            "state": self.state.name
        })

    def __eq__(self, __o: object) -> bool:
        if isinstance(__o, BugInfo):
            if __o.line_num == self.line_num and __o.bug_type == self.bug_type:
                return True
        return False

    def __hash__(self) -> int:
        return hash(self.line_num) + hash(self.bug_type)


class Report:
    """
    基础报告类，粒度为每个合约
    """

    def __init__(self, sol_path, contract_name, static_time_cost, gen_target_time_cost, dynamic_time_cost,
                 report_json_path, category):
        self.sol_path = sol_path  # 被测合约地址
        self.contract_name = contract_name  # 被测合约名称
        self.time_cost = static_time_cost + dynamic_time_cost  # 总时间消耗
        self.static_time_cost = static_time_cost  # 静态分析
        self.gen_target_time_cost = gen_target_time_cost  # 制导信息生成所用的时间
        self.dynamic_time_cost = dynamic_time_cost  # 符号执行所用的时间
        self.report_json_path = report_json_path  # 存储报告文件的地址.不同的检测工具.存储的地址不一样
        self.category = category
        self.bug_infos: List[BugInfo] = []  # 漏洞总集合
        self.tp: List[BugInfo] = []
        self.fp: List[BugInfo] = []
        self.fn: List[BugInfo] = []
        self.un: List[BugInfo] = []

    def append_bug_info(self, bug_info: BugInfo):
        self.bug_infos.append(bug_info)

    def check_with_smart_bugs_ground_truth(self):
        # 与基础类型对应的truth的地址.目前是smartbugs
        smart_bugs_ground_truth_file_path = '/root/smartbugs/dataset/vulnerabilities.json'
        truth = json.load(open(smart_bugs_ground_truth_file_path))
        find = False  # truth中是否能够查找到这被测文件.不能则代表该被测文件没有ground truth
        for one_truth in truth:
            # 查找被测文件的地址
            if one_truth['path'] == self.sol_path.replace("/root/smartbugs/dataset", "dataset").replace("/smart_bugs_Curated", ""):
                find = True  # 这个被测文件存在于truth里
                for vulnerability in one_truth['vulnerabilities']:
                    category = vulnerability['category']
                    for vulnerability_line in vulnerability['lines']:
                        mark = False  # 是否找到这个漏洞.没找到的漏洞.需要添加到FN里
                        for bug_info in self.bug_infos:
                            assert bug_info.bug_type.name == category.upper(), "应该指定类型进行匹配"
                            if bug_info.state == BugInfoState.FP:  # 先假定是误报.然后和smart_bugs_ground_truth中的bug进行比较.匹配成功就是TP
                                if vulnerability_line == bug_info.line_num and bug_info.bug_type == convert(category):
                                    bug_info.state = BugInfoState.TP
                                    mark = True
                                    break
                        if not mark:  # 没找到.加入到漏报FN中
                            un_find_bug_info = BugInfo(
                                vulnerability_line, convert(category), category)
                            un_find_bug_info.state = BugInfoState.FN
                            self.bug_infos.append(un_find_bug_info)
                break
        assert find, "truth中没有找到被测文件"  # 如果没找到.说明truth存在问题

    def count_tp_fp_fn_un(self):
        for bug_info in self.bug_infos:
            if bug_info.state == BugInfoState.TP:
                self.tp.append(bug_info)
            if bug_info.state == BugInfoState.FP:
                self.fp.append(bug_info)
            if bug_info.state == BugInfoState.FN:
                self.fn.append(bug_info)
            if bug_info.state == BugInfoState.UN:
                self.un.append(bug_info)
        assert len(self.tp) + len(self.fp) + len(self.fn) + \
               len(self.un) == len(self.bug_infos), "漏洞报告的数量不一致"

    def drop_duplication(self):
        self.bug_infos = list(set(self.bug_infos))


class MythrilReport(Report):
    """
    Mythril报告结果,粒度为每个合约
    """

    def __init__(self, sol_path, contract_name, static_time_cost, gen_target_time_cost, dynamic_time_cost,
                 report_json_path, report_type,
                 mode: Mode, coverage_txt_path, category,
                 bug_detect_mode: BugInfoDetectMode = BugInfoDetectMode.ONE_TIME, target_info_file_path=""):
        super().__init__(sol_path, contract_name, static_time_cost, gen_target_time_cost,
                         dynamic_time_cost, report_json_path, category)
        self.report_type = report_type
        self.mode: Mode = mode
        self.coverage_txt_path = coverage_txt_path
        self.init_coverage = 0
        self.runtime_coverage = 0
        self.process_report()
        self.drop_duplication()
        self.check_with_smart_bugs_ground_truth()
        self.count_tp_fp_fn_un()
        if bug_detect_mode == BugInfoDetectMode.EACH_TIME:
            # 如果是EACH模式, 那么这个成员变量表示, 是否这个单独目标被检测出来了, 默认为FN, 也就是没检测出来
            bug = BugInfo(-1, BugType.I_DO_NOT_KNOW, "this is origin bug, ready to check")
            bug.state = BugInfoState.FN
            self.is_origin_bug_checked = bug
            assert os.path.exists(target_info_file_path), "存放制导信息的文件不存在"
            self.target_info_file_path = target_info_file_path
            self.check_with_origin_bug_index()
        self.read_update_coverage()

    def check_with_origin_bug_index(self):
        """
        仅和自己的制导文件中, 标记为origin_bug的漏洞进行匹配 , 若检测出来了origin_bug, 那么就是tp, 其他的均为附赠品, 附赠品依旧遵循smart_bugs_ground_truth的逻辑继续检测
        """
        should_check_origin_bug = []
        target_info = json.load(open(self.target_info_file_path, "r"))
        for line_num, bugs in target_info.items():
            for bug in bugs:
                if bug['origin_bug']:
                    assert int(line_num) == int(bug['line_num']), "漏洞的key行数和漏洞信息内部的行数应该是一致的"
                    should_check_origin_bug.append(bug)
        assert len(should_check_origin_bug) == 1, "应该存在唯一一个origin漏洞"
        should_check_origin_bug = should_check_origin_bug[0]
        bug = BugInfo(int(should_check_origin_bug['line_num']), convert(should_check_origin_bug['bug_type']), should_check_origin_bug['bug_type'])
        bug.state = BugInfoState.FN  # 该指定的单一漏洞是需要被检测的, 默认设置为FN, 现在开始查看tp里是否有这个漏洞
        for tp in self.tp:
            if tp.line_num == bug.line_num:
                bug.state = BugInfoState.TP  # 这个单一漏洞确实被检测到了, 真好, 将其设置为TP
                break
        self.is_origin_bug_checked = bug

    def __str__(self):
        return 'MythrilReport: ' + self.sol_path + ':' + self.contract_name + '\n' \
               + "time: " + str(self.time_cost)[:6] + "s" + "\n\t\t" \
               + str(self.static_time_cost)[:5] + "s" + " + " + str(self.dynamic_time_cost)[:5] + "s" + '\n' \
               + "init_coverage: " + str(self.init_coverage) + "%" + '\n' \
               + "runtime_coverage: " + str(self.runtime_coverage) + "%" + '\n' \
               + "TP: " + str(self.tp) + '\n' \
               + "FP: " + str(self.fp) + '\n' \
               + "FN: " + str(self.fn) + '\n' \
               + "UN: " + str(self.un)

    def to_json_str(self):
        return json.dumps({
            "sol_path": self.sol_path,
            "contract_name": self.contract_name,
            "time_cost": self.time_cost,
            "static_time_cost": self.static_time_cost,
            "dynamic_time_cost": self.dynamic_time_cost,
            "report_type": self.report_type,
            "mode": self.mode.name,
            "coverage_txt_path": self.coverage_txt_path,
            "init_coverage": self.init_coverage,
            "runtime_coverage": self.runtime_coverage,
            "tp": [one.to_json_str() for one in self.tp],
            "fp": [one.to_json_str() for one in self.fp],
            "fn": [one.to_json_str() for one in self.fn],
            "un": [one.to_json_str() for one in self.un]
        })

    def process_report(self):
        if self.report_type == 'jsonv1':
            self.jsonv1()

    def jsonv1(self):
        mythril_json = json.load(open(self.report_json_path))
        assert mythril_json['error'] is None and mythril_json['success'], f"{self.mode}的Mythril的报告显示.执行存在错误"
        for issue in mythril_json['issues']:
            if "lineno" not in issue.keys() or "function" in issue['code']:  # 可能是函数
                continue
            line_num = issue['lineno']
            if issue['title'] == 'State access after external call':  # 忽略对状态变量的修改这一报告，无法被truth人工标注
                continue
            bug_type = convert(issue['title'])
            if bug_type.name != self.category.upper():  # 由于smartbugs仅标注单一漏洞, 因此其他类型的漏洞即使检测出来了, 也无法确认TP, 故将其跳过
                continue
            if bug_type == BugType.REENTRANCY:
                # 按照solidity分隔符对语句切分
                split_title = re.split('\.|\(|\)', issue['code'])
                if 'call' not in split_title:  # 如果call这个函数调用不在语句中, 说明是mythril找到的外部函数调用, 这类不被smartbugs标注,去掉
                    continue
            bug_info = BugInfo(line_num, bug_type, issue['title'])
            self.append_bug_info(bug_info)

    def jsonv2(self, mythril_json2):
        pass

    def read_update_coverage(self):
        with open(self.coverage_txt_path, 'r') as f:
            lines = f.readlines()
            if len(lines) != 2:
                print(f"警告！{self.mode}覆盖信息文件存在错误,覆盖文件存在{len(lines)}行")
                return
            self.init_coverage = float(
                lines[0].split(" ")[1].replace("%", "").strip())
            self.runtime_coverage = float(
                lines[1].split(" ")[1].replace("%", "").strip())


class SlitherReport(Report):
    """
    单个contract的Slither报告结果
    """

    def __init__(self, sol_path, contract_name, static_time_cost, target_gen_time_cost, dynamic_time_cost, report_json_path, category):
        super().__init__(sol_path, contract_name, static_time_cost, target_gen_time_cost, dynamic_time_cost, report_json_path, category)
        self.state_write_function_counter = 0
        self.state_write_modifier_counter = 0
        self.struct_write_counter = 0
        self.inner_func_call_counter = 0
        self.state_write_function_lines = set()
        self.state_write_modifier_lines = set()
        self.struct_write_lines = set()
        self.inner_func_call_lines = set()
        self.process_report()
        self.drop_duplication()
        self.check_with_smart_bugs_ground_truth()
        self.count_tp_fp_fn_un()
        self.correct()

    def process_report(self):
        slither_json = json.load(open(self.report_json_path))
        for line, infos in slither_json.items():
            for info in infos:
                assert int(line) == int(info['line_num'])
                line_num = int(line)
                if "written_target" in info['bug_type']:
                    if info['is_finded_in_src_map']:
                        if info['bug_type'] == 'written_target_function':
                            self.state_write_function_counter += 1
                            self.state_write_function_lines.add(line_num)
                        elif info['bug_type'] == "written_target_modifier":
                            self.state_write_modifier_counter += 1
                            self.state_write_modifier_lines.add(line_num)
                        elif info['bug_type'] == "written_target_function_and_modifier":
                            self.state_write_function_counter += 1
                            self.state_write_modifier_counter += 1
                            self.state_write_function_lines.add(line_num)
                            self.state_write_modifier_lines.add(line_num)
                        elif info['bug_type'] == 'written_target_struct_write':
                            self.struct_write_counter += 1
                            self.struct_write_lines.add(line_num)
                        elif info['bug_type'] == "written_target_inner_func_call_s":
                            self.inner_func_call_counter += 1
                            self.inner_func_call_lines.add(line_num)
                    continue
                if info['bug_type'] == "function_sig":
                    continue
                if info['slither_node_type'] == 'function':
                    continue
                bug_type = convert(info['bug_type'])
                if bug_type.name != self.category.upper():  # 在smartbugs指定的漏洞类别中, 只识别指定的漏洞
                    continue
                if bug_type == BugType.REENTRANCY:
                    underlying_type = info['slither_addtional_field'].get("underlying_type", 'empty')
                    if underlying_type == 'variables_written':  # 跳过slither中，重入漏洞的状态变量修改语句，这类语句不会被truth人工标注
                        continue
                bug_info = BugInfo(line_num, bug_type, info['bug_type'])
                self.append_bug_info(bug_info)

    def correct(self):
        self.state_write_function_counter = len(self.state_write_function_lines)
        self.state_write_modifier_counter = len(self.state_write_modifier_lines)
        self.struct_write_counter = len(self.struct_write_lines)
        self.inner_func_call_counter = len(self.inner_func_call_lines)

    def __str__(self):
        return 'SlitherReport: ' + self.sol_path + ':' + self.contract_name + '\n' \
               + "time: " + str(self.time_cost)[:6] + "s" + "\n\t\t" \
               + str(self.static_time_cost)[:5] + "s" + " + " + str(self.dynamic_time_cost)[:5] + "s" + '\n' \
               + "TP: " + str(self.tp) + '\n' \
               + "FP: " + str(self.fp) + '\n' \
               + "FN: " + str(self.fn) + '\n' \
               + "UN: " + str(self.un)


class CompareWithTwoBugInfoState(Enum):
    TP_TP = 1
    TP_FP = 2
    TP_FN = 3
    TP_UN = 4
    FP_TP = 5
    FP_FP = 6
    FP_FN = 7
    FP_UN = 8
    FN_TP = 9
    FN_FP = 10
    FN_FN = 11
    FN_UN = 12
    UN_TP = 13
    UN_FP = 14
    UN_FN = 15
    UN_UN = 16


class ReportFileLevel:
    """
    单一sol文件的整体报告类,粒度为文件级别
    实际上,这才是和truth能够匹配的等级,该等级应该由低等级的Report合成出来
    请注意,最后统计tp、fp、tn和un,都应该以文件级别作为标准,否则会出现slither重复统计tp,mythril多统计fn的情况
    """

    def __init__(self, sol_path, slither_time_cost, target_gen_time_cost, target_time_cost, fully_time_cost,
                 slither_r: List[SlitherReport],
                 mythril_target_r: List[MythrilReport],
                 mythril_fully_r: List[MythrilReport],
                 detect_mode: BugInfoDetectMode = BugInfoDetectMode.ONE_TIME) -> None:
        self.sol_path = sol_path
        self.slither_time_cost = slither_time_cost
        self.target_gen_time_cost = target_gen_time_cost
        self.target_time_cost = target_time_cost
        self.fully_time_cost = fully_time_cost
        # 通过together函数生成的,将多个合约报告融合,去除掉fn错误报告的情况
        self.slither_summary: ToolBugStateSummary = None
        self.mythril_target_summary: ToolBugStateSummary = None
        self.mythril_fully_summary: ToolBugStateSummary = None
        # mythril的两种模式与slither进行匹配的结果
        self.slither_with_mythril_target: MatchingWithSlitherMythril = None
        self.slither_with_mythril_fully: MatchingWithSlitherMythril = None
        # 以下信息不再具有参考意义,仅用来人工核对结果,下面的信息包含fn冗余、fn在tp中出现但是没删去的情况
        self.slither_r = slither_r
        self.mythril_target_r = mythril_target_r
        self.mythril_fully_r = mythril_fully_r
        self.together()  # 聚合各个合约的结果,考虑fn漏洞出现在tp中但没有被删除的情况
        self.check_match()  # 统计slither与两个模式mythril的匹配信息
        if detect_mode == BugInfoDetectMode.EACH_TIME:
            bug = BugInfo(-1, BugType.I_DO_NOT_KNOW, "this is origin bug, ready to check")
            bug.state = BugInfoState.FN
            self.is_origin_checked = bug  # 在target的EACH模式下, 结合多个合约来看, 是否检测出来的指定的那一个漏洞?
            # 和fully比起来, 在多合约情况下, 两者的结果是什么?
            self.compare_with_target_and_fully_in_each = CompareWithTwoBugInfoState.FN_FN
            self.check_origin_bug()

    def check_origin_bug(self):
        """
        将多个合约聚合, 一起查看是否检测出来了origin_bug
        """
        for r in self.mythril_target_r:
            # 即使是FN , 也需要把相关的信息给与together级别的数据类里
            self.is_origin_checked = r.is_origin_bug_checked
            if r.is_origin_bug_checked.state == BugInfoState.TP:
                self.compare_with_target_and_fully_in_each = CompareWithTwoBugInfoState.TP_FN
                break  # 此时保留了这个TP状态, 所以无需再次给is_origin_checked赋值了
        for t in self.mythril_fully_summary.tp:
            if t.line_num == self.is_origin_checked.line_num:
                if self.compare_with_target_and_fully_in_each == CompareWithTwoBugInfoState.TP_FN:
                    self.compare_with_target_and_fully_in_each = CompareWithTwoBugInfoState.TP_TP
                else:
                    self.compare_with_target_and_fully_in_each = CompareWithTwoBugInfoState.FN_TP

    def check_match(self):
        self.slither_with_mythril_target = MatchingWithSlitherMythril(self.slither_summary, self.mythril_target_summary)
        self.slither_with_mythril_fully = MatchingWithSlitherMythril(self.slither_summary, self.mythril_fully_summary)

    def together(self):
        assert len(self.slither_r) >= 1  # 至少含有一个报告
        assert len(self.slither_r) == len(self.mythril_target_r) == len(self.mythril_fully_r)  # 三个工具的报告数量应该一致,均为合约数量
        # slither的文件级别的报告, 所有的报告应该是相同的, 取第一个
        slither_report_one = self.slither_r[0]
        self.slither_summary = ToolBugStateSummary(slither_report_one.tp, slither_report_one.fp,
                                                   slither_report_one.fn, slither_report_one.un, Tool.SLITHER.name,
                                                   slither_report_one.fp)
        # 处理mythril_target
        temp_tp, temp_fp, temp_fn, temp_un = [], [], [], []
        for mythril_target_one in self.mythril_target_r:
            temp_tp.extend(mythril_target_one.tp)
            temp_fp.extend(mythril_target_one.fp)
            temp_un.extend(mythril_target_one.un)
            temp_fn.extend(mythril_target_one.fn)
        temp_tp, temp_fp, temp_fn, temp_un = set(temp_tp), set(temp_fp), set(temp_fn), set(temp_un)
        for checking_fn in temp_fn.copy():
            for matching_tp in temp_tp:
                if checking_fn.line_num == matching_tp.line_num and checking_fn.bug_type == checking_fn.bug_type:
                    # 该漏洞并不是没检测到(fn)，而是因为mythril逐个合约检测，该漏洞在其他合约中。因此若该漏洞被其他合约mythril检测到了，移除fn
                    temp_fn.remove(checking_fn)
        temp_tp, temp_fp, temp_fn, temp_un = list(temp_tp), list(temp_fp), list(temp_fn), list(temp_un)
        self.mythril_target_summary = ToolBugStateSummary(temp_tp, temp_fp, temp_fn, temp_un, Tool.MYTHRIL_TARGET.name, temp_fp)
        # 处理mythril_fully
        temp_tp, temp_fp, temp_fn, temp_un = [], [], [], []
        for mythril_fully_one in self.mythril_fully_r:
            temp_tp.extend(mythril_fully_one.tp)
            temp_fp.extend(mythril_fully_one.fp)
            temp_un.extend(mythril_fully_one.un)
            temp_fn.extend(mythril_fully_one.fn)
        temp_tp, temp_fp, temp_fn, temp_un = set(temp_tp), set(temp_fp), set(temp_fn), set(temp_un)
        for checking_fn in temp_fn.copy():
            for matching_tp in temp_tp:
                if checking_fn.line_num == matching_tp.line_num and checking_fn.bug_type == checking_fn.bug_type:
                    # 该漏洞并不是没检测到(fn)，而是因为mythril逐个合约检测，该漏洞在其他合约中。因此若该漏洞被其他合约mythril检测到了，移除fn
                    temp_fn.remove(checking_fn)
        temp_tp, temp_fp, temp_fn, temp_un = list(
            temp_tp), list(temp_fp), list(temp_fn), list(temp_un)
        self.mythril_fully_summary = ToolBugStateSummary(
            temp_tp, temp_fp, temp_fn, temp_un, Tool.MYTHRIL_TARGET.name, temp_fp)


class ToolBugStateSummary:
    """
    单一工具的漏洞检测情况

    """

    def __init__(self, tp, fp, fn, un, tool_name, fp_detail):
        self.tool_name = tool_name
        self.tp = tp
        self.fp = fp
        self.fn = fn
        self.un = un
        self.fp_detail = fp_detail


class MatchingWithSlitherMythril:
    """
    单一sol文件中,slither与mythril在不同模式下的匹配情况,总共有下面这些情况
       slither-tp         mythril-tp       【确认成功】
       slither-tp         mythril-fp        (不可能的情况,因为任何工具的tp都被truth确认过,只要tool的行号和基础类型匹配,一定是tp)
       slither-tp         mythril-unfind   【mythril能力问题】

       slither-fp         mythril-tp        (不可能的情况)
       slither-fp         mythril-fp       【俩工具的能力问题】
       slither-fp         mythril-unfind   【确认成功】

       slither-unfind     mythril-tp        【slither能力问题, 这是mythril的附加品】
       slither-unfind     mythril-fp        【反向确认? mythril的能力问题】
       slither-unfind     mythril-unfind    【无法确认的东西, 这个业务意义不存在, 因为必须以一个为标准】
    """

    def __init__(self, slither_summary: ToolBugStateSummary, mythril_summary: ToolBugStateSummary) -> None:
        self.tp_tp: List[(BugInfo, BugInfo)] = []
        self.tp_fp: List[BugInfo] = []
        self.tp_unfind: List[BugInfo] = []
        self.fp_tp: List[BugInfo] = []
        self.fp_fp: List[BugInfo] = []
        self.fp_unfind: List[BugInfo] = []
        self.unfind_tp: List[BugInfo] = []
        self.unfind_fp = []
        self.unfind_unfind = []
        self.check(slither_summary, mythril_summary)

    def check(self, slither_summary: ToolBugStateSummary, mythril_summary: ToolBugStateSummary):
        mythril_summary_tp_copy = mythril_summary.tp.copy()
        mythril_summary_fp_copy = mythril_summary.fp.copy()

        slither_summary_tp_copy = slither_summary.tp.copy()
        for slither_tp in slither_summary.tp:
            for mythril_tp in mythril_summary.tp:
                if slither_tp.line_num == mythril_tp.line_num and slither_tp.bug_type == mythril_tp.bug_type:
                    self.tp_tp.append((slither_tp, mythril_tp))
                    slither_summary_tp_copy.remove(slither_tp)
                    mythril_summary_tp_copy.remove(mythril_tp)
            for mythril_fp in mythril_summary.fp:
                if slither_tp.line_num == mythril_fp.line_num and slither_tp.bug_type == mythril_fp.bug_type:
                    raise "slither的tp被mythril认为是fp,这是不可能出现的情况,请确认"
        self.tp_unfind = slither_summary_tp_copy

        slither_summary_fp_copy = slither_summary.fp.copy()
        for slither_fp in slither_summary.fp:
            for mythril_tp in mythril_summary.tp:
                if slither_fp.line_num == mythril_tp.line_num and slither_fp.bug_type == mythril_tp.bug_type:
                    raise "mythril的tp被slither认为是fp,这是不可能出现的情况,请确认"
            for mythril_fp in mythril_summary.fp:
                if slither_fp.line_num == mythril_fp.line_num and slither_fp.bug_type == mythril_fp.bug_type:
                    self.fp_fp.append((slither_fp, mythril_fp))
                    slither_summary_fp_copy.remove(slither_fp)
                    mythril_summary_fp_copy.remove(mythril_fp)
        self.fp_unfind = slither_summary_fp_copy

        self.unfind_tp = mythril_summary_tp_copy
        self.unfind_fp = mythril_summary_fp_copy


class SmartTargetReport:
    """
    smart_target工具报告结果
    """

    def __init__(self, slither_report: SlitherReport, mythril_target_report: MythrilReport, mythtil_fully_report: MythrilReport, summary: Summary) -> None:
        self.summary = summary
        self.tool_summary = dict()
        self.slither_report = slither_report
        self.mythril_target_report = mythril_target_report
        self.mythtil_fully_report = mythtil_fully_report
        self.count_each_tool_bug_state()

    def count_each_tool_bug_state(self):
        self.tool_summary = {
            'slither': ToolBugStateSummary(len(self.slither_report.tp), len(self.slither_report.fp),
                                           len(self.slither_report.fn), len(
                    self.slither_report.un), "slither",
                                           self.slither_report.fp),
            'mythril_target': ToolBugStateSummary(len(self.mythril_target_report.tp),
                                                  len(self.mythril_target_report.fp),
                                                  len(self.mythril_target_report.fn),
                                                  len(self.mythril_target_report.un), "mythril_target",
                                                  self.mythril_target_report.fp),
            'mythril_fully': ToolBugStateSummary(len(self.mythtil_fully_report.tp), len(self.mythtil_fully_report.fp),
                                                 len(self.mythtil_fully_report.fn), len(
                    self.mythtil_fully_report.un),
                                                 "mythril_fully", self.mythtil_fully_report.fp),
        }


class SmartTargetSummaryReport:
    """
    smart-target工具总体误报、漏报情况统计
    """

    def __init__(self, detail, detect_mode: BugInfoDetectMode = BugInfoDetectMode.ONE_TIME) -> None:
        """
        detail是每个文件每个合约三个工具的具体报告
        detect_mode是目标设置的方式, 理论上除了RQ2都是一次性将所有漏洞都作为目标, RQ2是将每个漏洞单独设为目标

        """
        self.detect_mode = detect_mode
        self.summary = {
            'tp': 0,
            'fp': 0,
            'fn': 0,
            'un': 0
        }
        self.summary_each_tool = {
            Tool.SLITHER.name: {
                'tp': 0,
                'fp': 0,
                'fn': 0,
                'un': 0,
                'recall': 0,
                'precision': 0,
                'f1': 0,
                'tp_bug_type_counter': {},
                'fp_bug_type_counter': {},
                'fn_bug_type_counter': {}
            },
            Tool.MYTHRIL_TARGET.name: {
                'tp': 0,
                'fp': 0,
                'fn': 0,
                'un': 0,
                'recall': 0,
                'precision': 0,
                'f1': 0,
                'tp_bug_type_counter': {},
                'fp_bug_type_counter': {},
                'fn_bug_type_counter': {}
            },
            Tool.MYTHRIL_FULLY.name: {
                'tp': 0,
                'fp': 0,
                'fn': 0,
                'un': 0,
                'recall': 0,
                'precision': 0,
                'f1': 0,
                'tp_bug_type_counter': {},
                'fp_bug_type_counter': {},
                'fn_bug_type_counter': {}
            }
        }
        self.matching = {
            'slither_with_mythril_target': {
                'tp_tp': 0,
                'tp_fp': 0,
                'tp_unfind': 0,
                'fp_tp': 0,
                'fp_fp': 0,
                'fp_unfind': 0,
                'unfind_tp': 0,
                'unfind_fp': 0,
                'unfind_unfind': 0
            },
            'slither_with_mythril_fully': {
                'tp_tp': 0,
                'tp_fp': 0,
                'tp_unfind': 0,
                'fp_tp': 0,
                'fp_fp': 0,
                'fp_unfind': 0,
                'unfind_tp': 0,
                'unfind_fp': 0,
                'unfind_unfind': 0
            }
        }
        self.tp_set = {
            Tool.SLITHER.name: {},
            Tool.MYTHRIL_TARGET.name: {},
            Tool.MYTHRIL_FULLY.name: {}
        }
        self.reproduct_set = {
            'S & T': {},
            'S & F': {},
            'T & F': {}
        }
        self.reproduct_counter = {
            'S & T': 0,
            'S & F': 0,
            'T & F': 0
        }
        self.each_bug_type_summary: EachBugTypeReport = None
        # together的结构 [sol_path => {slither:[tp,fp,...]}, {mythril_target:[tp, ...]}]
        self.together: Dict[str, ReportFileLevel] = dict()
        self.detail: Dict[str, SmartTargetReport] = dict(sorted(detail.items(), key=lambda s: s[0]))

        self.together_each_sol_file()  # together是所有统计工作的基础, 必须将统计工作建立在文件级别
        self.gen_summary()
        self.gen_each_summary()
        self.count_fp_type()
        self.count_tp_type()
        self.count_fn_type()
        self.count_matching()
        self.gen_each_bug_type_summary()
        self.count_tp_set()

    def count_tp_set(self):
        for path, rlf in self.together.items():
            slither_s = rlf.slither_summary.tp
            mythril_target_s = rlf.mythril_target_summary.tp
            mythril_fully_s = rlf.mythril_fully_summary.tp
            for bug_type in [BugType.REENTRANCY, BugType.ACCESS_CONTROL, BugType.UNCHECKED_LOW_LEVEL_CALLS, BugType.TIME_MANIPULATION]:
                slither_tp_lines = [f'{path}-{s.line_num}' for s in slither_s if s.bug_type == bug_type]
                mythril_target_tp_lines = [f'{path}-{s.line_num}' for s in mythril_target_s if s.bug_type == bug_type]
                mythril_fully_tp_lines = [f'{path}-{s.line_num}' for s in mythril_fully_s if s.bug_type == bug_type]
                slither_tp_set = self.tp_set.get(Tool.SLITHER.name, {}).get(bug_type.name, set())
                mythril_target_tp_set = self.tp_set.get(Tool.MYTHRIL_TARGET.name, {}).get(bug_type.name, set())
                mythril_fully_tp_set = self.tp_set.get(Tool.MYTHRIL_FULLY.name, {}).get(bug_type.name, set())
                slither_tp_set.update(slither_tp_lines)
                mythril_target_tp_set.update(mythril_target_tp_lines)
                mythril_fully_tp_set.update(mythril_fully_tp_lines)
                self.tp_set[Tool.SLITHER.name][bug_type.name] = slither_tp_set
                self.tp_set[Tool.MYTHRIL_TARGET.name][bug_type.name] = mythril_target_tp_set
                self.tp_set[Tool.MYTHRIL_FULLY.name][bug_type.name] = mythril_fully_tp_set
        loguru.logger.info(f'{Tool.SLITHER.name} tp set: {len(self.tp_set[Tool.SLITHER.name])}')
        loguru.logger.info(f'{Tool.MYTHRIL_TARGET.name} tp set: {len(self.tp_set[Tool.MYTHRIL_TARGET.name])}')
        loguru.logger.info(f'{Tool.MYTHRIL_FULLY.name} tp set: {len(self.tp_set[Tool.MYTHRIL_FULLY.name])}')
        for bug_type in [BugType.REENTRANCY, BugType.ACCESS_CONTROL, BugType.UNCHECKED_LOW_LEVEL_CALLS, BugType.TIME_MANIPULATION]:
            self.reproduct_set['S & T'][bug_type.name] = self.tp_set[Tool.SLITHER.name][bug_type.name] & self.tp_set[Tool.MYTHRIL_TARGET.name][bug_type.name]
            self.reproduct_set['S & F'][bug_type.name] = self.tp_set[Tool.SLITHER.name][bug_type.name] & self.tp_set[Tool.MYTHRIL_FULLY.name][bug_type.name]
            self.reproduct_set['T & F'][bug_type.name] = self.tp_set[Tool.MYTHRIL_TARGET.name][bug_type.name] & self.tp_set[Tool.MYTHRIL_FULLY.name][bug_type.name]
            self.reproduct_counter['S & T'] += len(self.reproduct_set['S & T'][bug_type.name])
            self.reproduct_counter['S & F'] += len(self.reproduct_set['S & F'][bug_type.name])
            self.reproduct_counter['T & F'] += len(self.reproduct_set['T & F'][bug_type.name])
        loguru.logger.info(f'{Tool.SLITHER.name} & {Tool.MYTHRIL_TARGET.name} tp set: {len(self.reproduct_set["S & T"])}')
        loguru.logger.info(f'{Tool.SLITHER.name} & {Tool.MYTHRIL_FULLY.name} tp set: {len(self.reproduct_set["S & F"])}')
        loguru.logger.info(f'{Tool.MYTHRIL_TARGET.name} & {Tool.MYTHRIL_FULLY.name} tp set: {len(self.reproduct_set["T & F"])}')

    def gen_each_bug_type_summary(self):
        self.each_bug_type_summary = EachBugTypeReport(self.together)

    def count_matching(self):
        for _, file_level_report in self.together.items():
            slither_with_mythril_target = file_level_report.slither_with_mythril_target

            self.matching['slither_with_mythril_target']['tp_tp'] += len(
                slither_with_mythril_target.tp_tp)
            self.matching['slither_with_mythril_target']['tp_fp'] += len(
                slither_with_mythril_target.tp_fp)
            self.matching['slither_with_mythril_target']['tp_unfind'] += len(
                slither_with_mythril_target.tp_unfind)
            self.matching['slither_with_mythril_target']['fp_tp'] += len(
                slither_with_mythril_target.fp_tp)
            self.matching['slither_with_mythril_target']['fp_fp'] += len(
                slither_with_mythril_target.fp_fp)
            self.matching['slither_with_mythril_target']['fp_unfind'] += len(
                slither_with_mythril_target.fp_unfind)
            self.matching['slither_with_mythril_target']['unfind_tp'] += len(
                slither_with_mythril_target.unfind_tp)
            self.matching['slither_with_mythril_target']['unfind_fp'] += len(
                slither_with_mythril_target.unfind_fp)
            self.matching['slither_with_mythril_target']['unfind_unfind'] += len(
                slither_with_mythril_target.unfind_unfind)

            slither_with_mythril_fully = file_level_report.slither_with_mythril_fully
            self.matching['slither_with_mythril_fully']['tp_tp'] += len(
                slither_with_mythril_fully.tp_tp)
            self.matching['slither_with_mythril_fully']['tp_fp'] += len(
                slither_with_mythril_fully.tp_fp)
            self.matching['slither_with_mythril_fully']['tp_unfind'] += len(
                slither_with_mythril_fully.tp_unfind)
            self.matching['slither_with_mythril_fully']['fp_tp'] += len(
                slither_with_mythril_fully.fp_tp)
            self.matching['slither_with_mythril_fully']['fp_fp'] += len(
                slither_with_mythril_fully.fp_fp)
            self.matching['slither_with_mythril_fully']['fp_unfind'] += len(
                slither_with_mythril_fully.fp_unfind)
            self.matching['slither_with_mythril_fully']['unfind_tp'] += len(
                slither_with_mythril_fully.unfind_tp)
            self.matching['slither_with_mythril_fully']['unfind_fp'] += len(
                slither_with_mythril_fully.unfind_fp)
            self.matching['slither_with_mythril_fully']['unfind_unfind'] += len(
                slither_with_mythril_fully.unfind_unfind)

    def together_each_sol_file(self):
        slither_report_each_file = dict()
        mythril_target_report_each_file = dict()
        mythril_fully_report_each_file = dict()
        time_cost_each_file = dict()
        sol_paths = set()
        for sol_path_and_contract_name, report in self.detail.items():  # 首先将各个工具的报告, 根据文件名作为聚类, 将报告聚起来
            slither_r = report.slither_report
            mythril_target_r = report.mythril_target_report
            mythril_fully_r = report.mythtil_fully_report

            sol_path_and_contract_name_list = sol_path_and_contract_name.split(":")
            if self.detect_mode == BugInfoDetectMode.ONE_TIME:
                assert len(sol_path_and_contract_name_list) == 2
            elif self.detect_mode == BugInfoDetectMode.EACH_TIME:
                assert len(sol_path_and_contract_name_list) == 3
            sol_path = sol_path_and_contract_name_list[0]
            if self.detect_mode == BugInfoDetectMode.EACH_TIME:
                sol_path = sol_path + ":" + sol_path_and_contract_name_list[2]
            sol_paths.add(sol_path)

            temp = slither_report_each_file.get(sol_path, [])
            temp.append(slither_r)
            slither_report_each_file[sol_path] = temp

            temp = mythril_target_report_each_file.get(sol_path, [])
            temp.append(mythril_target_r)
            mythril_target_report_each_file[sol_path] = temp

            temp = mythril_fully_report_each_file.get(sol_path, [])
            temp.append(mythril_fully_r)
            mythril_fully_report_each_file[sol_path] = temp
            # 记录融合后的时间 ,  这里 slither时间应该都是相同的, 因为slither已整体文件作为分析
            # 生成制导信息的时间、mythril-target和mythril-fully的时间采用所有合约累加的形式
            if sol_path in time_cost_each_file.keys():
                time_cost_each_file[sol_path] = {
                    'slither_time': slither_r.static_time_cost,
                    'target_gen_time': report.summary.target_gen_time + time_cost_each_file[sol_path]['target_gen_time'],
                    'mythril_target_time': mythril_target_r.dynamic_time_cost + time_cost_each_file[sol_path]['mythril_target_time'],
                    'mythril_fully_time': mythril_fully_r.dynamic_time_cost + time_cost_each_file[sol_path]['mythril_fully_time']
                }
            else:
                time_cost_each_file[sol_path] = {
                    'slither_time': slither_r.static_time_cost,
                    'target_gen_time': report.summary.target_gen_time,
                    'mythril_target_time': mythril_target_r.dynamic_time_cost,
                    'mythril_fully_time': mythril_fully_r.dynamic_time_cost
                }
            assert slither_r.static_time_cost == report.summary.slither_time
            assert mythril_target_r.dynamic_time_cost == report.summary.mythril_target_time
            assert mythril_fully_r.dynamic_time_cost == report.summary.mythril_fully_time

        assert sol_paths == slither_report_each_file.keys()
        assert sol_paths == mythril_target_report_each_file.keys()
        assert sol_paths == mythril_fully_report_each_file.keys()
        assert sol_paths == time_cost_each_file.keys()
        for sol_path in sol_paths:  # 将各个文件的各个工具产生的报告集合, 放入ReportFileLevel这个class, 这个类的内部会进行聚合
            self.together[sol_path] = ReportFileLevel(sol_path,
                                                      time_cost_each_file[sol_path]['slither_time'],
                                                      time_cost_each_file[sol_path]['target_gen_time'],
                                                      time_cost_each_file[sol_path]['mythril_target_time'],
                                                      time_cost_each_file[sol_path]['mythril_fully_time'],
                                                      slither_report_each_file[sol_path],
                                                      mythril_target_report_each_file[sol_path],
                                                      mythril_fully_report_each_file[sol_path],
                                                      self.detect_mode)
        self.together = dict(sorted(self.together.items(), key=lambda s: s[0]))

    def gen_summary(self):
        for _, report in self.together.items():
            slither_r = report.slither_summary
            mythril_target_r = report.mythril_target_summary
            mythril_fully_r = report.mythril_fully_summary
            self.summary['tp'] += len(slither_r.tp) + \
                                  len(mythril_target_r.tp) + len(mythril_fully_r.tp)
            self.summary['fp'] += len(slither_r.fp) + \
                                  len(mythril_target_r.fp) + len(mythril_fully_r.fp)
            self.summary['fn'] += len(slither_r.fn) + \
                                  len(mythril_target_r.fn) + len(mythril_fully_r.fn)
            self.summary['un'] += len(slither_r.un) + \
                                  len(mythril_target_r.un) + len(mythril_fully_r.un)

    def count_fn_type(self):
        slither_fn_counter = 0  # 记录该工具的总fn数量，用于确认代码正确性
        mythril_target_fn_counter = 0
        mythril_fully_fn_counter = 0
        for bug_type in BugType:  # 先初始化字典，让基础漏洞的fn数量均为0
            self.summary_each_tool[Tool.SLITHER.name]['fn_bug_type_counter'][bug_type.name] = 0
            self.summary_each_tool[Tool.MYTHRIL_TARGET.name]['fn_bug_type_counter'][bug_type.name] = 0
            self.summary_each_tool[Tool.MYTHRIL_FULLY.name]['fn_bug_type_counter'][bug_type.name] = 0
        for _, report in self.together.items():
            slither_r = report.slither_summary
            mythril_target_r = report.mythril_target_summary
            mythril_fully_r = report.mythril_fully_summary
            for fn_r in slither_r.fn:  # slither的fn
                # 根据该fn漏洞的漏洞类型，使字典中的数据更新
                self.summary_each_tool[Tool.SLITHER.name]['fn_bug_type_counter'][fn_r.bug_type.name] += 1
                slither_fn_counter += 1
            for fn_r in mythril_target_r.fn:
                self.summary_each_tool[Tool.MYTHRIL_TARGET.name]['fn_bug_type_counter'][fn_r.bug_type.name] += 1
                mythril_target_fn_counter += 1
            for fn_r in mythril_fully_r.fn:
                self.summary_each_tool[Tool.MYTHRIL_FULLY.name]['fn_bug_type_counter'][fn_r.bug_type.name] += 1
                mythril_fully_fn_counter += 1
        # 根据fn的数量，对基础漏洞类型排序
        self.summary_each_tool[Tool.SLITHER.name]['fn_bug_type_counter'] = dict(
            sorted(self.summary_each_tool[Tool.SLITHER.name]
                   ['fn_bug_type_counter'].items(), key=lambda s: s[1], reverse=True)
        )
        self.summary_each_tool[Tool.MYTHRIL_TARGET.name]['fn_bug_type_counter'] = dict(
            sorted(self.summary_each_tool[Tool.MYTHRIL_TARGET.name]
                   ['fn_bug_type_counter'].items(), key=lambda s: s[1], reverse=True)
        )
        self.summary_each_tool[Tool.MYTHRIL_FULLY.name]['fn_bug_type_counter'] = dict(
            sorted(self.summary_each_tool[Tool.MYTHRIL_FULLY.name]
                   ['fn_bug_type_counter'].items(), key=lambda s: s[1], reverse=True)
        )
        assert slither_fn_counter == self.summary_each_tool[Tool.SLITHER.name]['fn']
        assert mythril_target_fn_counter == self.summary_each_tool[Tool.MYTHRIL_TARGET.name]['fn']
        assert mythril_fully_fn_counter == self.summary_each_tool[Tool.MYTHRIL_FULLY.name]['fn']

    def count_fp_type(self):
        slither_fp_counter = 0  # 记录该工具的总fp数量，用于确认代码正确性
        mythril_target_fp_counter = 0
        mythril_fully_fp_counter = 0
        for bug_type in BugType:  # 先初始化字典，让基础漏洞的fp数量均为0
            self.summary_each_tool[Tool.SLITHER.name]['fp_bug_type_counter'][bug_type.name] = 0
            self.summary_each_tool[Tool.MYTHRIL_TARGET.name]['fp_bug_type_counter'][bug_type.name] = 0
            self.summary_each_tool[Tool.MYTHRIL_FULLY.name]['fp_bug_type_counter'][bug_type.name] = 0
        for _, report in self.together.items():
            slither_r = report.slither_summary
            mythril_target_r = report.mythril_target_summary
            mythril_fully_r = report.mythril_fully_summary
            for fp_r in slither_r.fp:  # slither的fp
                # 根据该fp漏洞的漏洞类型，使字典中的数据更新
                self.summary_each_tool[Tool.SLITHER.name]['fp_bug_type_counter'][fp_r.bug_type.name] += 1
                slither_fp_counter += 1
            for fp_r in mythril_target_r.fp:
                self.summary_each_tool[Tool.MYTHRIL_TARGET.name]['fp_bug_type_counter'][fp_r.bug_type.name] += 1
                mythril_target_fp_counter += 1
            for fp_r in mythril_fully_r.fp:
                self.summary_each_tool[Tool.MYTHRIL_FULLY.name]['fp_bug_type_counter'][fp_r.bug_type.name] += 1
                mythril_fully_fp_counter += 1
        # 根据fp的数量，对基础漏洞类型排序
        self.summary_each_tool[Tool.SLITHER.name]['fp_bug_type_counter'] = dict(
            sorted(self.summary_each_tool[Tool.SLITHER.name]
                   ['fp_bug_type_counter'].items(), key=lambda s: s[1], reverse=True)
        )
        self.summary_each_tool[Tool.MYTHRIL_TARGET.name]['fp_bug_type_counter'] = dict(
            sorted(self.summary_each_tool[Tool.MYTHRIL_TARGET.name]
                   ['fp_bug_type_counter'].items(), key=lambda s: s[1], reverse=True)
        )
        self.summary_each_tool[Tool.MYTHRIL_FULLY.name]['fp_bug_type_counter'] = dict(
            sorted(self.summary_each_tool[Tool.MYTHRIL_FULLY.name]
                   ['fp_bug_type_counter'].items(), key=lambda s: s[1], reverse=True)
        )
        assert slither_fp_counter == self.summary_each_tool[Tool.SLITHER.name]['fp']
        assert mythril_target_fp_counter == self.summary_each_tool[Tool.MYTHRIL_TARGET.name]['fp']
        assert mythril_fully_fp_counter == self.summary_each_tool[Tool.MYTHRIL_FULLY.name]['fp']

    def count_tp_type(self):
        slither_tp_counter = 0  # 记录该工具的总tp数量，用于确认代码正确性
        mythril_target_tp_counter = 0
        mythril_fully_tp_counter = 0
        for bug_type in BugType:  # 先初始化字典，让基础漏洞的tp数量均为0
            self.summary_each_tool[Tool.SLITHER.name]['tp_bug_type_counter'][bug_type.name] = 0
            self.summary_each_tool[Tool.MYTHRIL_TARGET.name]['tp_bug_type_counter'][bug_type.name] = 0
            self.summary_each_tool[Tool.MYTHRIL_FULLY.name]['tp_bug_type_counter'][bug_type.name] = 0
        for _, report in self.together.items():
            slither_r = report.slither_summary
            mythril_target_r = report.mythril_target_summary
            mythril_fully_r = report.mythril_fully_summary
            for tp_r in slither_r.tp:  # slither的tp
                # 根据该tp漏洞的漏洞类型，使字典中的数据更新
                self.summary_each_tool[Tool.SLITHER.name]['tp_bug_type_counter'][tp_r.bug_type.name] += 1
                slither_tp_counter += 1
            for tp_r in mythril_target_r.tp:
                self.summary_each_tool[Tool.MYTHRIL_TARGET.name]['tp_bug_type_counter'][tp_r.bug_type.name] += 1
                mythril_target_tp_counter += 1
            for tp_r in mythril_fully_r.tp:
                self.summary_each_tool[Tool.MYTHRIL_FULLY.name]['tp_bug_type_counter'][tp_r.bug_type.name] += 1
                mythril_fully_tp_counter += 1
        # 根据tp的数量，对基础漏洞类型排序
        self.summary_each_tool[Tool.SLITHER.name]['tp_bug_type_counter'] = dict(
            sorted(self.summary_each_tool[Tool.SLITHER.name]
                   ['tp_bug_type_counter'].items(), key=lambda s: s[1], reverse=True)
        )
        self.summary_each_tool[Tool.MYTHRIL_TARGET.name]['tp_bug_type_counter'] = dict(
            sorted(self.summary_each_tool[Tool.MYTHRIL_TARGET.name]
                   ['tp_bug_type_counter'].items(), key=lambda s: s[1], reverse=True)
        )
        self.summary_each_tool[Tool.MYTHRIL_FULLY.name]['tp_bug_type_counter'] = dict(
            sorted(self.summary_each_tool[Tool.MYTHRIL_FULLY.name]
                   ['tp_bug_type_counter'].items(), key=lambda s: s[1], reverse=True)
        )
        assert slither_tp_counter == self.summary_each_tool[Tool.SLITHER.name]['tp']
        assert mythril_target_tp_counter == self.summary_each_tool[Tool.MYTHRIL_TARGET.name]['tp']
        assert mythril_fully_tp_counter == self.summary_each_tool[Tool.MYTHRIL_FULLY.name]['tp']

    def gen_each_summary(self):
        for _, report in self.together.items():
            slither_r = report.slither_summary
            mythril_target_r = report.mythril_target_summary
            mythril_fully_r = report.mythril_fully_summary
            self.summary_each_tool[Tool.SLITHER.name]['tp'] += len(
                slither_r.tp)
            self.summary_each_tool[Tool.SLITHER.name]['fp'] += len(
                slither_r.fp)
            self.summary_each_tool[Tool.SLITHER.name]['fn'] += len(
                slither_r.fn)
            self.summary_each_tool[Tool.SLITHER.name]['un'] += len(
                slither_r.un)
            self.summary_each_tool[Tool.MYTHRIL_TARGET.name]['tp'] += len(
                mythril_target_r.tp)
            self.summary_each_tool[Tool.MYTHRIL_TARGET.name]['fp'] += len(
                mythril_target_r.fp)
            self.summary_each_tool[Tool.MYTHRIL_TARGET.name]['fn'] += len(
                mythril_target_r.fn)
            self.summary_each_tool[Tool.MYTHRIL_TARGET.name]['un'] += len(
                mythril_target_r.un)
            self.summary_each_tool[Tool.MYTHRIL_FULLY.name]['tp'] += len(
                mythril_fully_r.tp)
            self.summary_each_tool[Tool.MYTHRIL_FULLY.name]['fp'] += len(
                mythril_fully_r.fp)
            self.summary_each_tool[Tool.MYTHRIL_FULLY.name]['fn'] += len(
                mythril_fully_r.fn)
            self.summary_each_tool[Tool.MYTHRIL_FULLY.name]['un'] += len(
                mythril_fully_r.un)
        self.summary_each_tool[Tool.SLITHER.name]['recall'] = recall(
            self.summary_each_tool[Tool.SLITHER.name]['tp'], self.summary_each_tool[Tool.SLITHER.name]['fn'])
        self.summary_each_tool[Tool.SLITHER.name]['precision'] = recall(
            self.summary_each_tool[Tool.SLITHER.name]['tp'], self.summary_each_tool[Tool.SLITHER.name]['fp'])
        self.summary_each_tool[Tool.MYTHRIL_TARGET.name]['recall'] = recall(
            self.summary_each_tool[Tool.MYTHRIL_TARGET.name]['tp'],
            self.summary_each_tool[Tool.MYTHRIL_TARGET.name]['fn'])
        self.summary_each_tool[Tool.MYTHRIL_TARGET.name]['precision'] = recall(
            self.summary_each_tool[Tool.MYTHRIL_TARGET.name]['tp'],
            self.summary_each_tool[Tool.MYTHRIL_TARGET.name]['fp'])
        self.summary_each_tool[Tool.MYTHRIL_FULLY.name]['recall'] = recall(
            self.summary_each_tool[Tool.MYTHRIL_FULLY.name]['tp'],
            self.summary_each_tool[Tool.MYTHRIL_FULLY.name]['fn'])
        self.summary_each_tool[Tool.MYTHRIL_FULLY.name]['precision'] = recall(
            self.summary_each_tool[Tool.MYTHRIL_FULLY.name]['tp'],
            self.summary_each_tool[Tool.MYTHRIL_FULLY.name]['fp'])
        self.summary_each_tool[Tool.SLITHER.name]['f1'] = f1(
            self.summary_each_tool[Tool.SLITHER.name]['recall'], self.summary_each_tool[Tool.SLITHER.name]['precision'])
        self.summary_each_tool[Tool.MYTHRIL_TARGET.name]['f1'] = f1(
            self.summary_each_tool[Tool.MYTHRIL_TARGET.name]['recall'],
            self.summary_each_tool[Tool.MYTHRIL_TARGET.name]['precision'])
        self.summary_each_tool[Tool.MYTHRIL_FULLY.name]['f1'] = f1(
            self.summary_each_tool[Tool.MYTHRIL_FULLY.name]['recall'],
            self.summary_each_tool[Tool.MYTHRIL_FULLY.name]['precision'])


def json_outputer(obj):
    if isinstance(obj, Enum):  # 枚举类型没有dict.使用他们的name作为value
        return obj.name
    if isinstance(obj, set):
        return list(obj)
    return obj.__dict__


def recall(tp, fn):
    if tp == 0 and fn == 0:
        return 0
    return tp / (tp + fn)


def pricsion(tp, fp):
    if tp == 0 and fp == 0:
        return 0
    return tp / (tp + fp)


def f1(recall, precision):
    if recall == 0 and precision == 0:
        return 0
    return (2 * recall * precision) / (recall + precision)


class EachBugTypeBaseReport:
    def __init__(self, tool, bug_type) -> None:
        self.tool = tool
        self.bug_type = bug_type
        self.total_file_size = 0
        self.total_report_size = 0
        self.time_cost = 0
        self.tp_size = 0
        self.fp_size = 0
        self.fn_size = 0
        self.un_size = 0


class EachBugTypeReport:
    def __init__(self, together_summary) -> None:
        self.each_bug_type: Dict = {
            Tool.SLITHER.name: {
                'total_file_size': 0,
                'total_report_size': 0,
                'total_time_cost': 0,
                'total_tp_size': 0,
                'total_fp_size': 0,
                'total_fn_size': 0,
                'total_un_size': 0

            },
            Tool.MYTHRIL_TARGET.name: {
                'total_file_size': 0,
                'total_report_size': 0,
                'total_time_cost': 0,
                'total_tp_size': 0,
                'total_fp_size': 0,
                'total_fn_size': 0,
                'total_un_size': 0
            },
            Tool.MYTHRIL_FULLY.name: {
                'total_file_size': 0,
                'total_report_size': 0,
                'total_time_cost': 0,
                'total_tp_size': 0,
                'total_fp_size': 0,
                'total_fn_size': 0,
                'total_un_size': 0
            }
        }
        self.process(together_summary)

    def process(self, together_summary: Dict[str, ReportFileLevel]):
        for tool in Tool:
            for bug_type in BugType:
                self.each_bug_type[tool.name][bug_type.name] = EachBugTypeBaseReport(
                    tool.name, bug_type.name)
        for sol_path, file_level_r in together_summary.items():
            bug_type = sol_path.split("/")[-2]

            each_bug_type_base_report_each_tool = self.each_bug_type[Tool.SLITHER.name][bug_type.upper(
            )]
            summary = file_level_r.slither_summary
            each_bug_type_base_report_each_tool.total_file_size += 1
            tp = summary.tp
            fp = summary.fp
            fn = summary.fn
            un = summary.un
            # 报告出来的漏洞数量, 等于tp+fp+fn
            each_bug_type_base_report_each_tool.total_report_size += (
                    len(tp) + len(fp) + len(un))
            each_bug_type_base_report_each_tool.tp_size += len(tp)
            each_bug_type_base_report_each_tool.fp_size += len(fp)
            each_bug_type_base_report_each_tool.fn_size += len(fn)
            each_bug_type_base_report_each_tool.un_size += len(un)
            each_bug_type_base_report_each_tool.time_cost += file_level_r.slither_time_cost

            each_bug_type_base_report_each_tool = self.each_bug_type[Tool.MYTHRIL_TARGET.name][bug_type.upper(
            )]
            summary = file_level_r.mythril_target_summary
            each_bug_type_base_report_each_tool.total_file_size += 1
            tp = summary.tp
            fp = summary.fp
            fn = summary.fn
            un = summary.un
            each_bug_type_base_report_each_tool.total_report_size += (
                    len(tp) + len(fp) + len(un))
            each_bug_type_base_report_each_tool.tp_size += len(tp)
            each_bug_type_base_report_each_tool.fp_size += len(fp)
            each_bug_type_base_report_each_tool.fn_size += len(fn)
            each_bug_type_base_report_each_tool.un_size += len(un)
            each_bug_type_base_report_each_tool.time_cost += file_level_r.target_time_cost

            each_bug_type_base_report_each_tool = self.each_bug_type[Tool.MYTHRIL_FULLY.name][bug_type.upper(
            )]
            summary = file_level_r.mythril_fully_summary
            each_bug_type_base_report_each_tool.total_file_size += 1
            tp = summary.tp
            fp = summary.fp
            fn = summary.fn
            un = summary.un
            each_bug_type_base_report_each_tool.total_report_size += (
                    len(tp) + len(fp) + len(un))
            each_bug_type_base_report_each_tool.tp_size += len(tp)
            each_bug_type_base_report_each_tool.fp_size += len(fp)
            each_bug_type_base_report_each_tool.fn_size += len(fn)
            each_bug_type_base_report_each_tool.un_size += len(un)
            each_bug_type_base_report_each_tool.time_cost += file_level_r.fully_time_cost
        for tool in Tool:
            for bug_type in BugType:
                self.each_bug_type[tool.name]['total_file_size'] += self.each_bug_type[tool.name][
                    bug_type.name].total_file_size
                self.each_bug_type[tool.name]['total_report_size'] += self.each_bug_type[tool.name][
                    bug_type.name].total_report_size
                self.each_bug_type[tool.name]['total_time_cost'] += self.each_bug_type[tool.name][
                    bug_type.name].time_cost
                self.each_bug_type[tool.name]['total_tp_size'] += self.each_bug_type[tool.name][bug_type.name].tp_size
                self.each_bug_type[tool.name]['total_fp_size'] += self.each_bug_type[tool.name][bug_type.name].fp_size
                self.each_bug_type[tool.name]['total_fn_size'] += self.each_bug_type[tool.name][bug_type.name].fn_size
                self.each_bug_type[tool.name]['total_un_size'] += self.each_bug_type[tool.name][bug_type.name].un_size
            assert self.each_bug_type[tool.name]['total_report_size'] == self.each_bug_type[tool.name][
                'total_tp_size'] + \
                   self.each_bug_type[tool.name]['total_fp_size'] + \
                   self.each_bug_type[tool.name]['total_un_size']


class RQ2:
    def __init__(self, file_path, bug_index, target_mode_results, fully_mode_results, target_cost_time, fully_cost_time,
                 compart_with_target_and_fully) -> None:
        self.file_name = file_path
        self.bug_index = bug_index
        self.file_loc = len(open(file_path, "r").readlines())
        # self.bytecode_size=get_bytecode_size_from_path(file_path)
        self.bytecode_size = 0
        self.target_cost_time = target_cost_time
        self.fully_cost_time = fully_cost_time
        self.diff = fully_cost_time - target_cost_time
        self.compare_with_target_and_fully = compart_with_target_and_fully
        self.tp_tp = []
        self.tp_fp = []
        self.tp_unfind = []
        self.fp_tp = []
        self.fp_fp = []
        self.fp_unfind = []
        self.unfind_tp = []
        self.unfind_fp = []
        self.unfind_unfind = []
        self.analysis_mapping(target_mode_results, fully_mode_results)

    def analysis_mapping(self, target_mode_results, fully_mode_result):
        checking = MatchingWithSlitherMythril(target_mode_results, fully_mode_result)
        self.tp_tp = checking.tp_tp  # 越高越好
        self.tp_fp = checking.tp_fp  # 不可能
        self.tp_unfind = checking.tp_unfind  # 理论上不可能, 除非制导能检测出全量不能检测的漏洞
        self.fp_tp = checking.fp_tp  # 不可能
        self.fp_fp = checking.fp_fp  # 全量和制导都误报了, 越低越好
        # 理论上不可能, 这是说制导报告了, 但是是误报, 且全量的没报出来, 这是不好的情况, 说明制导导致部分限制条件被忽略了
        self.fp_unfind = checking.fp_unfind
        # 越低越好, 这个最好为0, 这表明制导报不出来的漏洞被全量报出来了, 说明制导的信息不够
        self.unfind_tp = checking.unfind_tp
        self.unfind_fp = checking.unfind_fp  # 越高越好, 但不太可能. 这说明全量存在误报, 但是在制导模式被消除了
        self.unfind_unfind = checking.unfind_unfind  # 不可能的情况
        assert len(self.tp_fp) == len(self.fp_tp) == len(self.unfind_unfind) == 0


class EmpricalResearchAndRQ1:
    """
    实证研究和RQ1的数据统计
    实证研究, 取Slither、Mythril-Fully和Manticore
    RQ1, 取Mythril-Fully和Mythril-Target
    """

    def __init__(self, bug_type, files_size, tool_name, reports_size, time_cost, tp_size, fp_size, fn_size,
                 un_size) -> None:
        self.bug_type = bug_type
        self.files_size = files_size
        self.total_bug_inner = tp_size + fn_size
        self.tool_name = tool_name
        self.reports_size = reports_size  # 该工具对于该漏洞类型, 报告的数量, 也就是tp+fp
        self.time_cost = time_cost
        self.tp_size = tp_size
        self.fp_size = fp_size
        self.fn_size = fn_size  # 误报数量
        assert un_size == 0
        assert reports_size == tp_size + fp_size
        self.recall = round(recall(tp_size, fn_size), 3)
        self.precision = round(pricsion(tp_size, fp_size), 3)
        self.f1 = round(f1(self.recall, self.precision), 3)

    def update(self):
        self.total_bug_inner = self.tp_size + self.fn_size
        assert self.reports_size == self.tp_size + self.fp_size
        self.recall = round(recall(self.tp_size, self.fn_size), 3)
        self.precision = round(pricsion(self.tp_size, self.fp_size), 3)
        self.f1 = round(f1(self.recall, self.precision), 3)


def gen_each_type_map_each_tool_tp_fp_fn_un_results_by_tool(smart_target_summary: SmartTargetSummaryReport, selected_tools,
                                                            skip_bug_type=[]):
    """
    生成以基础漏洞类型为第一列, 包含工具名, tp等信息的数据

    :param selected_tools: 只输出这个参数里面的工具
    :skip_bug_type: 跳过这个参数里面的基础漏洞

    """
    results = []
    last_row_sum_field = {}
    for tool_selected in selected_tools:
        last_row_sum_field[tool_selected] = EmpricalResearchAndRQ1('SUM_RESULT', 0, tool_selected, 0, 0, 0, 0, 0, 0)
    each_bug_type_reprot = smart_target_summary.each_bug_type_summary
    for tool_name, report in each_bug_type_reprot.each_bug_type.items():
        if tool_name not in selected_tools:
            continue
        for bug_type in BugType:
            if bug_type in skip_bug_type:
                continue
            bug_type_each_tool_summary = report[bug_type.name]
            total_file_size = bug_type_each_tool_summary.total_file_size
            tp_size = bug_type_each_tool_summary.tp_size
            fp_size = bug_type_each_tool_summary.fp_size
            fn_size = bug_type_each_tool_summary.fn_size
            un_size = bug_type_each_tool_summary.un_size
            # 对于特定的漏洞进行truth确认时, 不应该出现un的情况, 要么检测出来的tp, 要么就是没检测出来fn, 要么就是在别的位置检测出来的fp
            assert un_size == 0
            report_size = tp_size + fp_size + un_size
            # 纯粹的这个工具的时间, 对于制导mythril,为制导符号执行的时间, 不包括信息生成的时间
            time_cost = round(bug_type_each_tool_summary.time_cost, 3)
            if bug_type == BugType.I_DO_NOT_KNOW or bug_type == BugType.SKIP:
                # 这两种类型不是smartbugs的基础类型, 不可能存在tp等情况
                assert tp_size == fp_size == fn_size == un_size == time_cost == 0
                continue
            result = EmpricalResearchAndRQ1(bug_type.name, total_file_size, tool_name,
                                            report_size, time_cost, tp_size, fp_size, fn_size, un_size)
            # 用于统计最后一行, 全部漏洞的总体情况
            last_row_sum_field[tool_name].files_size += total_file_size
            last_row_sum_field[tool_name].reports_size += report_size
            last_row_sum_field[tool_name].time_cost += time_cost
            last_row_sum_field[tool_name].tp_size += tp_size
            last_row_sum_field[tool_name].fp_size += fp_size
            last_row_sum_field[tool_name].fn_size += fn_size
            results.append(result)
    for tool_selected in selected_tools:
        last_row_sum_field[tool_selected].update()
        results.append(last_row_sum_field[tool_selected])
    return results


def gen_time_cost_by_each_file(smart_target_summary: SmartTargetSummaryReport, skip_bug_type=[]):
    results = []
    for file_path, report in smart_target_summary.together.items():
        assert len(file_path.split(":")) == 2
        file_path, bug_index = file_path.split(":")[0], file_path.split(":")[1]
        file_path_split_list = file_path.split("/")
        assert len(file_path_split_list) == 6 and file_path_split_list[0] == '' and file_path_split_list[-1].endswith(".sol")
        category = file_path_split_list[-2]
        if category.upper() in [bug_type.name for bug_type in skip_bug_type]:
            continue
        rq2_r = RQ2(file_path, bug_index, report.mythril_target_summary, report.mythril_fully_summary, report.target_time_cost, report.fully_time_cost, report.compare_with_target_and_fully_in_each)
        results.append(rq2_r)
    return results


class RQ3():
    def __init__(self, bug_type, file_path, bug_index, state_function_write, state_modifier_write, struct_write, inner_fun_call) -> None:
        self.bug_type = bug_type
        self.state_function_write = state_function_write
        self.state_modifier_write = state_modifier_write
        self.struct_write = struct_write
        self.inner_fun_call = inner_fun_call
        self.file_path = file_path
        self.bug_index = bug_index


def gen_state_write_info(smart_target_summary: SmartTargetSummaryReport, skip_bug_type=[]):
    results = []
    for file_path, report in smart_target_summary.together.items():
        assert len(file_path.split(":")) == 2
        file_path, bug_index = file_path.split(":")[0], file_path.split(":")[1]
        file_path_split_list = file_path.split("/")
        assert len(file_path_split_list) == 6 and file_path_split_list[0] == '' and file_path_split_list[-1].endswith(".sol")
        category = file_path_split_list[-2]
        if category.upper() in [bug_type.name for bug_type in skip_bug_type]:
            continue
        state_function_write, state_modifier_write, struct_write, inner_fun_call = 0, 0, 0, 0
        for s_r in report.slither_r:
            state_function_write += s_r.state_write_function_counter
            state_modifier_write += s_r.state_write_modifier_counter
            struct_write += s_r.struct_write_counter
            inner_fun_call += s_r.inner_func_call_counter
        rq3_r = RQ3(category, file_path, bug_index, state_function_write, state_modifier_write, struct_write, inner_fun_call)
        results.append(rq3_r)
    return results


class Goal(Enum):
    EMPIRICAL_RESEARCH = 1
    RQ1 = 2
    RQ1_WITHOUT_NOT_SUPPORT = 3
    RQ2 = 4
    RQ3 = 5
    RQ2_EACH_BUG_TYPE = 6
    RQ3_WITH_STATE_WRITE_INFO = 7
    RQ2_REPRODUCTION = 8


def gen_results(smart_target_summary: SmartTargetSummaryReport, goal: Goal):
    if goal == Goal.EMPIRICAL_RESEARCH:
        return gen_each_type_map_each_tool_tp_fp_fn_un_results_by_tool(smart_target_summary,
                                                                       [Tool.SLITHER.name, Tool.MYTHRIL_FULLY.name],
                                                                       [BugType.ARITHMETIC, BugType.BAD_RANDOMNESS,
                                                                        BugType.DENIAL_OF_SERVICE,
                                                                        BugType.FRONT_RUNNING, BugType.OTHER,
                                                                        BugType.SHORT_ADDRESSES])
    elif goal == Goal.RQ1:
        return gen_each_type_map_each_tool_tp_fp_fn_un_results_by_tool(smart_target_summary, [Tool.MYTHRIL_FULLY.name,
                                                                                              Tool.MYTHRIL_TARGET.name])
    elif goal == Goal.RQ1_WITHOUT_NOT_SUPPORT:
        return gen_each_type_map_each_tool_tp_fp_fn_un_results_by_tool(smart_target_summary,
                                                                       [Tool.MYTHRIL_FULLY.name, Tool.MYTHRIL_TARGET.name],
                                                                       [BugType.ARITHMETIC, BugType.BAD_RANDOMNESS,
                                                                        BugType.DENIAL_OF_SERVICE,
                                                                        BugType.FRONT_RUNNING, BugType.OTHER,
                                                                        BugType.SHORT_ADDRESSES])
    elif goal == Goal.RQ2:
        return gen_time_cost_by_each_file(smart_target_summary, [BugType.BAD_RANDOMNESS, BugType.DENIAL_OF_SERVICE,
                                                                 BugType.FRONT_RUNNING, BugType.OTHER,
                                                                 BugType.SHORT_ADDRESSES])
    elif goal == Goal.RQ3:
        return gen_time_cost_by_each_file(smart_target_summary, [BugType.BAD_RANDOMNESS, BugType.DENIAL_OF_SERVICE,
                                                                 BugType.FRONT_RUNNING, BugType.OTHER,
                                                                 BugType.SHORT_ADDRESSES])
    elif goal == Goal.RQ2_EACH_BUG_TYPE:
        return gen_each_type_map_each_tool_tp_fp_fn_un_results_by_tool(smart_target_summary,
                                                                       [Tool.MYTHRIL_FULLY.name, Tool.MYTHRIL_TARGET.name],
                                                                       [BugType.ARITHMETIC, BugType.BAD_RANDOMNESS,
                                                                        BugType.DENIAL_OF_SERVICE,
                                                                        BugType.FRONT_RUNNING, BugType.OTHER,
                                                                        BugType.SHORT_ADDRESSES])
    elif goal == Goal.RQ3_WITH_STATE_WRITE_INFO:
        return gen_state_write_info(smart_target_summary, [BugType.BAD_RANDOMNESS, BugType.DENIAL_OF_SERVICE,
                                                           BugType.FRONT_RUNNING, BugType.OTHER,
                                                           BugType.SHORT_ADDRESSES])


def handle_empirical_research_and_rq1_df_to_excel(df, excel_path):
    df['tool_name'].replace(
        Tool.MYTHRIL_TARGET.name, 'Smart-Target', inplace=True)
    df['tool_name'].replace(
        Tool.MYTHRIL_FULLY.name, 'Mythril', inplace=True)
    df['tool_name'].replace(
        Tool.SLITHER.name, 'Slither', inplace=True)
    df['bug_type'].replace(BugType.ACCESS_CONTROL.name, "权限控制", inplace=True)
    df['bug_type'].replace(BugType.REENTRANCY.name, "重入", inplace=True)
    df['bug_type'].replace(BugType.TIME_MANIPULATION.name, "时间操控", inplace=True)
    df['bug_type'].replace(BugType.UNCHECKED_LOW_LEVEL_CALLS.name, "未检查低级调用", inplace=True)
    df['bug_type'].replace("SUM_RESULT", "总结", inplace=True)
    df.sort_values(["bug_type", "tool_name"], inplace=True, ascending=False)
    df.rename(columns={'bug_type': '漏洞类型', 'files_size': '文件数', 'total_bug_inner': '漏洞数', 'tool_name': '工具',
                       'reports_size': '报告数', 'time_cost': '时间', 'tp_size': 'TP', 'fp_size': 'FP',
                       'fn_size': 'FN',
                       'recall': 'Recall', 'precision': 'Precision', 'f1': 'F1'}, inplace=True)

    df.to_excel(excel_path, index=False, encoding="utf_8_sig")
    excel = load_workbook(excel_path)
    sheet = excel['Sheet1']
    for i in range(2, sheet.max_row, 2):
        assert sheet[i][1].value == sheet[i + 1][1].value
        sheet.merge_cells(start_row=i, end_row=i + 1,
                          start_column=1, end_column=1)
        assert sheet[i][2].value == sheet[i][2].value
        sheet.merge_cells(start_row=i, end_row=i + 1,
                          start_column=2, end_column=2)
        assert sheet[i][3].value == sheet[i][3].value
        sheet.merge_cells(start_row=i, end_row=i + 1,
                          start_column=3, end_column=3)
    excel.save(excel_path)


def handle_rq2_df_to_excel(df, excel_path, is_rq3=False):
    for index, row in df.iterrows():
        row['tp_tp'] = len(row['tp_tp'])
        row['tp_fp'] = len(row['tp_fp'])
        row['tp_unfind'] = len(row['tp_unfind'])
        row['fp_tp'] = len(row['fp_tp'])
        row['fp_fp'] = len(row['fp_fp'])
        row['fp_unfind'] = len(row['fp_unfind'])
        row['unfind_tp'] = len(row['unfind_tp'])
        row['unfind_fp'] = len(row['unfind_fp'])
        row['unfind_unfind'] = len(row['unfind_unfind'])
        df.loc[index] = row

    df.sort_values(["diff"], inplace=True)
    df.rename(columns={'file_name': '文件名', 'file_loc': '文件行数',
                       'target_cost_time': 'Smart-Target', 'fully_cost_time': 'Mythril',
                       'diff': 'Mythril比Smart-Target慢的时间', 'byte_code': '字节码总大小'}, inplace=True)
    if is_rq3:
        df.rename(columns={'Smart-Target': 'Smart-Target', 'Mythril': 'Smart-Target*',
                           'Mythril比Smart-Target慢的时间': 'Smart-Target*比Smart-Target慢的时间'}, inplace=True)

    df.to_excel(excel_path, index=False, encoding="utf_8_sig")
    df.to_csv(excel_path.replace(".xlsx", ".csv"), index=False, encoding="utf_8_sig")


def handle_rq2_each_bug_type_df_to_excel(df, excel_path):
    handle_empirical_research_and_rq1_df_to_excel(df, excel_path)


class RQ2BugReproduction():
    def __init__(self, bug_type, total_bug_count, tool_name, time_cost, reproduct_count) -> None:
        self.bug_type = bug_type
        self.total_bug_count = total_bug_count
        self.tool_name = tool_name
        self.time_cost = time_cost
        self.reproduct_count = reproduct_count
        self.reproduct_rate = 0

    def update(self):
        self.reproduct_rate = self.reproduct_count / self.total_bug_count


@loguru.logger.catch()
def handle_rq2_bug_reproduction(result: List[Dict]):
    """
    RQ2, 漏洞再现的相关数据
    """
    raw_data = []
    for bt in [BugType.ACCESS_CONTROL, BugType.ARITHMETIC, BugType.REENTRANCY, BugType.TIME_MANIPULATION, BugType.UNCHECKED_LOW_LEVEL_CALLS]:
        for tn in [Tool.MYTHRIL_FULLY, Tool.MYTHRIL_TARGET]:
            rq2 = RQ2BugReproduction(bt, 0, tn, 0, 0)
            raw_data.append(rq2)
    for r in result:
        file_path = r['file_name']
        assert len(file_path.split("/")) == 6 and ".sol" in file_path.split("/")[-1]
        category = convert(file_path.split("/")[-2])
        assert category in [BugType.ACCESS_CONTROL, BugType.ARITHMETIC, BugType.REENTRANCY, BugType.TIME_MANIPULATION, BugType.UNCHECKED_LOW_LEVEL_CALLS]
        compare_result = r['compare_with_target_and_fully']
        assert compare_result in [CompareWithTwoBugInfoState.TP_TP, CompareWithTwoBugInfoState.FN_FN]
        is_fully_product = False
        is_target_product = False
        if compare_result == CompareWithTwoBugInfoState.TP_TP:
            is_fully_product = True
            is_target_product = True
        elif compare_result == CompareWithTwoBugInfoState.FN_FN:
            is_fully_product = False
            is_target_product = False
        else:
            raise Exception("RQ2中, 目标制导没有达到和全量符号执行一样的效果")
        for rq2 in raw_data:
            if rq2.bug_type == category and rq2.tool_name == Tool.MYTHRIL_FULLY:
                rq2.total_bug_count += 1
                rq2.time_cost += r['fully_cost_time']
                if is_fully_product:
                    rq2.reproduct_count += 1
            elif rq2.bug_type == category and rq2.tool_name == Tool.MYTHRIL_TARGET:
                rq2.total_bug_count += 1
                rq2.time_cost += r['target_cost_time']
                if is_target_product:
                    rq2.reproduct_count += 1
    for rq2 in raw_data:
        rq2.update()
    return raw_data


def handle_rq3_result_to_plot(result: List[Dict], base_path, empirical_result):
    """
    将RQ3的结果绘图
    需要注意的是, result里的fully是smart-target*的结果
    """
    bug_compare_counter = {}
    time_compare_counter = {}
    for o in result:
        file_name = o['file_name']
        bug_index = o['bug_index']
        target_cost_time = o['target_cost_time']
        target_star_cost_time = o['fully_cost_time']
        compare_result = o['compare_with_target_and_fully']
        assert compare_result in [CompareWithTwoBugInfoState.TP_TP, CompareWithTwoBugInfoState.TP_FN, CompareWithTwoBugInfoState.FN_TP, CompareWithTwoBugInfoState.FN_FN]
        assert len(file_name.split("/")) == 6 and ".sol" in file_name.split("/")[-1]
        category = convert(file_name.split("/")[-2])
        assert category in [BugType.ACCESS_CONTROL, BugType.ARITHMETIC, BugType.REENTRANCY, BugType.TIME_MANIPULATION, BugType.UNCHECKED_LOW_LEVEL_CALLS]
        tp_target_star = bug_compare_counter.get(category, {}).get('tp_target_star', 0)
        tp_target = bug_compare_counter.get(category, {}).get('tp_target', 0)
        fn_target_star = bug_compare_counter.get(category, {}).get('fn_target_star', 0)
        fn_target = bug_compare_counter.get(category, {}).get('fn_target', 0)
        if compare_result == CompareWithTwoBugInfoState.TP_TP:
            tp_target_star += 1
            tp_target += 1
        elif compare_result == CompareWithTwoBugInfoState.TP_FN:
            tp_target += 1
            fn_target_star += 1
        elif compare_result == CompareWithTwoBugInfoState.FN_TP:
            loguru.logger.warning("存在开启前置操作分析后造成FN的情况,请检查")
        elif compare_result == CompareWithTwoBugInfoState.FN_FN:
            fn_target += 1
            fn_target_star += 1
        bug_compare_counter[category] = {'tp_target_star': tp_target_star, 'tp_target': tp_target, 'fn_target_star': fn_target_star, 'fn_target': fn_target}
        target_cost_time_cumulate = time_compare_counter.get(category, {}).get('target_cost_time', 0)
        target_star_cost_time_cumulate = time_compare_counter.get(category, {}).get('target_star_cost_time', 0)
        time_compare_counter[category] = {'target_cost_time': target_cost_time_cumulate + target_cost_time, 'target_star_cost_time': target_star_cost_time_cumulate + target_star_cost_time}

    df_result_bug_compare = []
    for bug_type, detail in bug_compare_counter.items():
        fully_tp = -1
        fully_fn = -1
        for o in empirical_result:
            if o['bug_type'] == bug_type.name and o['tool_name'] == Tool.MYTHRIL_FULLY.name:
                fully_tp = o['tp_size']
                fully_fn = o['fn_size']
                break
        assert fully_tp != -1 and fully_fn != -1
        tp = detail['tp_target']
        tp_star = detail['tp_target_star']
        fn = detail['fn_target']
        fn_star = detail['fn_target_star']
        df_result_bug_compare.append({
            'bug_type': bug_type.name,
            "method": u"Smart-Target$^{-}$",
            "tp": tp_star,
            "fn": fn_star,
        })
        df_result_bug_compare.append({
            'bug_type': bug_type.name,
            "method": "Smart-Target",
            "tp": tp,
            "fn": fn,
        })
        # df_result_bug_compare.append({
        #     'bug_type': bug_type.name,
        #     "method": "Mythril",
        #     "tp": fully_tp,
        #     "fn": fully_fn,
        # })
    df_result_bug_compare.sort(key=lambda x: x['tp'])
    bug_compare_counter_df = pandas.DataFrame(df_result_bug_compare)
    bug_compare_counter_df['bug_type'].replace(BugType.ACCESS_CONTROL.name, "AC", inplace=True)
    bug_compare_counter_df['bug_type'].replace(BugType.ARITHMETIC.name, "ARM", inplace=True)
    bug_compare_counter_df['bug_type'].replace(BugType.REENTRANCY.name, "RE", inplace=True)
    bug_compare_counter_df['bug_type'].replace(BugType.TIME_MANIPULATION.name, "TM", inplace=True)
    bug_compare_counter_df['bug_type'].replace(BugType.UNCHECKED_LOW_LEVEL_CALLS.name, "ULLC", inplace=True)
    bug_compare_counter_df.sort_values(by=['method'], inplace=True, ascending=False)
    bug_compare_counter_df.to_excel(os.path.join(base_path, f"{Goal.RQ3}.bug_compare.xlsx"), index=True, encoding="utf_8_sig")
    df_result_time_compare = []
    for bug_type, detail in time_compare_counter.items():
        fully_cost_time = -1
        for o in empirical_result:
            if o['bug_type'] == bug_type.name and o['tool_name'] == Tool.MYTHRIL_FULLY.name:
                fully_cost_time = o['time_cost']
                break
        assert fully_cost_time != -1
        target_cost_time = detail['target_cost_time']
        target_star_cost_time = detail['target_star_cost_time']
        df_result_time_compare.append({
            'bug_type': bug_type.name,
            "method": u"Smart-Target$^{-}$",
            "cost_time": target_star_cost_time,
        })
        df_result_time_compare.append({
            'bug_type': bug_type.name,
            "method": "Smart-Target",
            "cost_time": target_cost_time,
        })
        # df_result_time_compare.append({
        #     'bug_type': bug_type.name,
        #     "method": "Mythril",
        #     "cost_time": fully_cost_time,
        # })
    df_result_time_compare.sort(key=lambda x: x['cost_time'])
    time_compare_counter_df = pandas.DataFrame(df_result_time_compare)
    time_compare_counter_df['bug_type'].replace(BugType.ACCESS_CONTROL.name, "AC", inplace=True)
    time_compare_counter_df['bug_type'].replace(BugType.ARITHMETIC.name, "ARM", inplace=True)
    time_compare_counter_df['bug_type'].replace(BugType.REENTRANCY.name, "RE", inplace=True)
    time_compare_counter_df['bug_type'].replace(BugType.TIME_MANIPULATION.name, "TM", inplace=True)
    time_compare_counter_df['bug_type'].replace(BugType.UNCHECKED_LOW_LEVEL_CALLS.name, "ULLC", inplace=True)
    time_compare_counter_df.sort_values(by=['method'], inplace=True, ascending=False)
    time_compare_counter_df.to_excel(os.path.join(base_path, f"{Goal.RQ3}.time_compare.xlsx"), index=True, encoding="utf_8_sig")

    # 将结果绘图
    # 先绘制bug类型的比较
    plt.figure(figsize=(5, 7))
    seaborn.barplot(x="bug_type", y="tp", data=bug_compare_counter_df, hue="method")
    plt.xlabel("Bug Type")
    plt.ylabel("True Positive")
    plt.legend(loc="best")
    plt.savefig(os.path.join(base_path, f"{Goal.RQ3}.bug_compare.png"), dpi=500, bbox_inches="tight")

    # 再绘制时间比较
    plt.figure(figsize=(5, 7))
    seaborn.barplot(x="bug_type", y="cost_time", data=time_compare_counter_df, hue="method")
    plt.xlabel("Bug Type")
    plt.ylabel("Cost Time")
    plt.legend(loc="best")
    plt.savefig(os.path.join(base_path, f"{Goal.RQ3}.time_compare.png"), dpi=500, bbox_inches="tight")

    # 组合为两个子图
    plt.figure(figsize=(10, 7))
    plt.subplot(1, 2, 1)
    seaborn.barplot(x="bug_type", y="tp", data=bug_compare_counter_df, hue="method")
    plt.xlabel("漏洞类型\n(a)安全漏洞检测数量对比")
    plt.ylabel("检测出漏洞的数量")
    plt.legend(loc="best")
    plt.subplot(1, 2, 2)
    seaborn.barplot(x="bug_type", y="cost_time", data=time_compare_counter_df, hue="method")
    plt.xlabel("漏洞类型\n(b)检测时间消耗对比")
    plt.ylabel("检测时间")

    plt.legend(loc="best")
    plt.subplots_adjust(wspace=0.5)
    plt.savefig(os.path.join(base_path, f"{Goal.RQ3}.bug_time_compare.png"), dpi=500, bbox_inches="tight")


def handle_rq2_df_to_plot(df, plot_path, bar_path, rq2_results, base_path):
    # 折线图部分
    # 由于被重命名了, 这里取的名字应该与handle_rq2_df_to_excel里的一样
    y1 = df['Smart-Target'].values
    # y1 = [1500 if y >= 1500 else y for y in y1]
    y2 = df['Mythril'].values
    # y2 = [1500 if y >= 1500 else y for y in y2]
    plt.plot(y1, color='blue')
    plt.plot(y2, color='red')
    plt.savefig(plot_path, dpi=500, bbox_inches='tight')
    plt.close()
    # 柱状图部分
    multiple_data = {
        0: 0,
        1: 0,
        2: 0,
        3: 0,
        4: 0,
        5: 0,
        6: 0,
        7: 0,
        8: 0,
        9: 0,
        10: 0,
        11: 0,
        12: 0,
        13: 0,
        14: 0,
        15: 0
    }
    y1 = df['Smart-Target'].values
    y2 = df['Mythril'].values
    assert len(y1) == len(y2)
    for i in range(len(y1)):
        multiple = int(y2[i] / y1[i])
        if multiple >= 15:
            multiple_data[15] += 1
        else:
            multiple_data[multiple] += 1
    plt.bar(multiple_data.keys(), multiple_data.values())
    plt.savefig(bar_path, dpi=500, bbox_inches='tight')
    print(multiple_data)
    # 双柱状图部分
    data = []
    for m in [Tool.MYTHRIL_FULLY, Tool.MYTHRIL_TARGET]:
        for time_zone in [10, 30, 60, 120, 180, 240, 300, 600, 1200, 1800, 3600]:
            temp_dict = {
                'method': m.name,
                'time_zone': time_zone,
                'tp': 0
            }
            data.append(temp_dict)
    for index, row in df.iterrows():
        if row['compare_with_target_and_fully'] == CompareWithTwoBugInfoState.TP_TP:
            for d in data:
                if d['method'] == Tool.MYTHRIL_TARGET.name and row['Smart-Target'] <= d['time_zone']:
                    d['tp'] += 1
                elif d['method'] == Tool.MYTHRIL_FULLY.name and row['Mythril'] <= d['time_zone']:
                    d['tp'] += 1
    for d in data:
        if d['time_zone'] < 60:
            d['time_zone'] = str(d['time_zone']) + 's'
        else:
            d['time_zone'] = str(d['time_zone'] // 60) + 'm'
    df_each_time_zone = pandas.DataFrame(data)
    df_each_time_zone['method'].replace(Tool.MYTHRIL_TARGET.name, "Smart-Target", inplace=True)
    df_each_time_zone['method'].replace(Tool.MYTHRIL_FULLY.name, "Mythril", inplace=True)
    df_each_time_zone.to_csv(os.path.join(base_path, f"{Goal.RQ2}.each_time_zone.csv"), index=False)
    plt.figure(figsize=(10, 7))
    seaborn.barplot(x="time_zone", y="tp", data=df_each_time_zone, hue="method")
    plt.xlabel("Time")
    plt.ylabel("Reproducibility")
    plt.legend(loc="best")
    plt.savefig(os.path.join(base_path, f"{Goal.RQ2}.bug_time_compare_bar.png"), dpi=500, bbox_inches="tight")

    # 双折线图, 每个折线图是柱状图的值
    plt.figure(figsize=(10, 7))
    data_liner = []
    for m in [Tool.MYTHRIL_FULLY, Tool.MYTHRIL_TARGET]:
        for time_zone in range(0, 3600, 60):
            temp_dict = {
                'method': m.name,
                'time_zone': time_zone,
                'tp': 0
            }
            data_liner.append(temp_dict)
    for index, row in df.iterrows():
        if row['compare_with_target_and_fully'] == CompareWithTwoBugInfoState.TP_TP:
            for d in data_liner:
                if d['method'] == Tool.MYTHRIL_TARGET.name and row['Smart-Target'] <= d['time_zone']:
                    d['tp'] += 1
                elif d['method'] == Tool.MYTHRIL_FULLY.name and row['Mythril'] <= d['time_zone']:
                    d['tp'] += 1
    df_data_liner = pandas.DataFrame(data_liner)
    df_data_liner['method'].replace(Tool.MYTHRIL_TARGET.name, "Smart-Target", inplace=True)
    df_data_liner['method'].replace(Tool.MYTHRIL_FULLY.name, "Mythril", inplace=True)
    # version1 不带有到时间轴的延长线
    plt.figure(figsize=(10, 7))
    seaborn.lineplot(x="time_zone", y="tp", data=df_data_liner, hue="method", marker="o")

    plt.xlabel("Time")
    plt.ylabel("Reproducibility")
    plt.legend(loc="best")

    top_y_size = max(df_data_liner['tp'].values) + 5
    plt.ylim(-5, top_y_size)
    plt.axvline(x=120, ymin=(49 + 5) / (top_y_size + 5), ymax=(85 + 5) / (top_y_size + 5), color='g', linestyle='dashed')

    plt.xlim(-60, 3660)
    plt.axhline(y=76, xmin=(120 + 60) / (3660 + 60), xmax=(1200 + 60) / (3660 + 60), color='r', linestyle='dashed')
    plt.xticks(range(0, 3601, 300))

    plt.savefig(os.path.join(base_path, f"{Goal.RQ2}.bug_time_compare_liner.png"), dpi=500, bbox_inches="tight")

    # version2 带有到时间轴的延长线
    plt.figure(figsize=(10, 7))
    seaborn.lineplot(x="time_zone", y="tp", data=df_data_liner, hue="method", marker="o")

    plt.xlabel("复现时间")
    plt.ylabel("复现漏洞数量")
    plt.legend(loc="best")

    top_y_size = max(df_data_liner['tp'].values) + 5
    plt.ylim(-5, top_y_size)
    plt.axvline(x=120, ymin=(49 + 5) / (top_y_size + 5), ymax=(85 + 5) / (top_y_size + 5), color='g', linestyle='dashed')
    plt.axvline(x=120, ymin=0, ymax=(85 + 5) / (top_y_size + 5), color='g', linestyle='dashed')

    plt.xlim(-60, 3660)
    # plt.axhline(y=76,xmin=(120+60)/(3660+60),xmax=(1200+60)/(3660+60),color='r',linestyle='dashed')
    plt.axvline(x=1200, ymin=0, ymax=(76 + 5) / (top_y_size + 5), color='r', linestyle='dashed')
    plt.xticks(range(0, 3601, 300))

    plt.savefig(os.path.join(base_path, f"{Goal.RQ2}.bug_time_compare_liner_with_timestamp.png"), dpi=500, bbox_inches="tight")

    # 双折线图部分, 取柱状图的差值
    data_sub = []
    for m in [Tool.MYTHRIL_FULLY, Tool.MYTHRIL_TARGET]:
        for time_zone in range(0, 3600, 60):
            time_zone_tuple = (time_zone, time_zone + 60)
            temp_dict = {
                'method': m.name,
                'time_zone': time_zone_tuple,
                'tp': 0
            }
            data_sub.append(temp_dict)
    for index, row in df.iterrows():
        if row['compare_with_target_and_fully'] == CompareWithTwoBugInfoState.TP_TP:
            for d in data_sub:
                if d['method'] == Tool.MYTHRIL_TARGET.name and row['Smart-Target'] >= d['time_zone'][0] and row['Smart-Target'] < d['time_zone'][1]:
                    d['tp'] += 1
                elif d['method'] == Tool.MYTHRIL_FULLY.name and row['Mythril'] >= d['time_zone'][0] and row['Mythril'] < d['time_zone'][1]:
                    d['tp'] += 1
    for d in data_sub:
        d['time_zone'] = int(d['time_zone'][1])
        if d['time_zone'] < 60:
            d['time_zone'] = str(d['time_zone']) + 's'
        else:
            d['time_zone'] = str(d['time_zone'] // 60) + 'm'
    df_each_time_zone_sub = pandas.DataFrame(data_sub)
    df_each_time_zone_sub['method'].replace(Tool.MYTHRIL_TARGET.name, "Smart-Target", inplace=True)
    df_each_time_zone_sub['method'].replace(Tool.MYTHRIL_FULLY.name, "Mythril", inplace=True)
    plt.figure(figsize=(10, 7))
    seaborn.lineplot(x="time_zone", y="tp", data=df_each_time_zone_sub, hue="method", marker="o")
    plt.xlabel("Time")
    plt.ylabel("Reproducibility")
    plt.legend(loc="best")
    # plt.xticks(rotation=60)
    plt.savefig(os.path.join(base_path, f"{Goal.RQ2}.bug_time_compare_sub.png"), dpi=500, bbox_inches="tight")


def get_bytecode_size_from_path(path):
    res = 0
    cy = CryticCompile(target=path)
    for name, bytecode in cy.compilation_units[path].bytecodes_runtime.items():
        res += len(bytecode)
    return res


def create_cfg(_runtime_bytecode):
    total_bb, linked_bb = 0, 0
    total_ins, linked_ins = 0, 0
    try:
        cfg = cfg_builder.CFG(_runtime_bytecode)
        total_bb = len(cfg.basic_blocks)
        visited = set()

        def deep(bb):
            if bb in visited:
                return
            visited.add(bb)
            for succ in bb.all_outgoing_basic_blocks:
                deep(succ)

        deep(cfg.entry_point)
        linked_bb = len(visited)

        def calu_ins_size(bbs):
            res = 0
            for bb in bbs:
                res += len(bb.instructions)
            return res

        total_ins = calu_ins_size(cfg.basic_blocks)
        linked_ins = calu_ins_size(visited)
    except BaseException as e:
        cfg = None
    return cfg, total_bb, linked_bb, total_ins, linked_ins


def verify_data_complete_and_threat_analysis(dataset_dir, empirical, rq1, rq2, rq3, verify_result_path):
    not_include = []
    for category, file, contract, reason in SKIP_SOL_CONTRACT_NAME_PAIR:
        not_include.append(category + ":" + file + ":" + contract)
    assert len(set(not_include)) == len(not_include)
    datasets = []
    cfg_success = []
    cfg_fail = []
    compile_fail = []
    sol_files = []
    total_bb_sum, linked_bb_sum = 0, 0
    total_ins_sum, linked_ins_sum = 0, 0
    for root, dirs, files in os.walk(dataset_dir):
        for file in files:
            if file.endswith(".sol"):
                path = os.path.join(root, file)
                category = path.split("/")[-2]
                sol_files.append(path)
                try:
                    cy = CryticCompile(target=path)
                    for i in cy.compilation_units[path].contracts_names_without_libraries:
                        bytecode = cy.compilation_units[path].bytecodes_runtime[i]
                        cfg, total_bb, linked_bb, total_ins, linked_ins = create_cfg(bytecode)
                        if cfg is None:
                            cfg_fail.append(category + ":" + file.replace(".sol", "") + ":" + i)
                            loguru.logger.error(f"{category}:{file.replace('.sol', '')}:{i}的CFG创建失败")
                        else:
                            cfg_success.append(category + ":" + file.replace(".sol", "") + ":" + i)
                            total_bb_sum += total_bb
                            linked_bb_sum += linked_bb
                            total_ins_sum += total_ins
                            linked_ins_sum += linked_ins
                            # if linked_bb/total_bb <0.8:
                            #     print(f"{category} {file} {i} {linked_bb/total_bb:.2%} {linked_ins/total_ins:.2%}")
                        if category + ":" + file.replace(".sol", "") + ":" + i in not_include:
                            continue
                        datasets.append(category + ":" + file.replace(".sol", "") + ":" + i)
                except BaseException as e:
                    loguru.logger.error(f"{path}编译错误")
                    compile_fail.append(category + ":" + file.replace(".sol", ""))
    assert len(set(datasets)) == len(datasets)
    datasets = set(datasets)

    empirical_dirs = []
    for root, dirs, files in os.walk(empirical):
        for dir in dirs:
            if ":" in dir:
                category = root.split("/")[-1]
                empirical_dirs.append(category + ":" + dir)

    assert len(set(empirical_dirs)) == len(empirical_dirs)
    empirical_dirs = set(empirical_dirs)
    print(f"EM比All少了{datasets - empirical_dirs}")

    rq1_dirs = []
    for root, dirs, files in os.walk(rq1):
        for dir in dirs:
            if ":" in dir:
                category = root.split("/")[-1]
                rq1_dirs.append(category + ":" + dir)

    assert len(set(rq1_dirs)) == len(rq1_dirs)
    rq1_dirs = set(rq1_dirs)
    print(f"rq1比All少了{datasets - rq1_dirs}")

    rq2_dirs = []
    for root, dirs, files in os.walk(rq2):
        for dir in dirs:
            if ":" in dir:
                category = root.split("/")[-1]
                rq2_dirs.append(category + ":" + dir)

    assert len(set(rq2_dirs)) == len(rq2_dirs)
    rq2_dirs = set(rq2_dirs)
    print(f"rq2比All少了{datasets - rq2_dirs}")

    rq3_dirs = []
    for root, dirs, files in os.walk(rq3):
        for dir in dirs:
            if ":" in dir:
                category = root.split("/")[-1]
                rq3_dirs.append(category + ":" + dir)

    assert len(set(rq3_dirs)) == len(rq3_dirs)
    rq3_dirs = set(rq3_dirs)
    print(f"rq3比All少了{datasets - rq3_dirs}")
    print(f"cfg success {len(cfg_success)}")
    print(f"cfg fail {len(cfg_fail)}")
    print(f"compile fail {len(compile_fail)}")
    print(f"dataset contract {len(datasets)}")
    print(f"sol files {len(sol_files)}")
    print(f"total bb {total_bb_sum}")
    print(f"linked bb {linked_bb_sum}")
    print(f"total ins {total_ins_sum}")
    print(f"linked ins {linked_ins_sum}")
    with open(os.path.join(verify_result_path, "verify_result.json"), "w") as f:
        json.dump({"cfg_success": cfg_success, "cfg_fail": cfg_fail, "compile_fail": compile_fail}, f)
    bug_infos = json.load(open(os.path.join(dataset_dir, "vulnerabilities.json")))
    bug_info_counter = {}
    for bug_info in bug_infos:
        name = bug_info["name"]
        path = bug_info["path"]
        vulnerabilities = bug_info["vulnerabilities"]
        category = path.split("/")[-2]
        for v in vulnerabilities:
            lines = v["lines"]
            assert category == v["category"]
            count = bug_info_counter.get(category, 0)
            count += len(lines)
            bug_info_counter[category] = count
    print(f"bug info counter {bug_info_counter}")


def draw_venn_plot_rq1(rq1_r):
    """
    rq1的韦恩图
    """
    pass


if __name__ == "__main__":
    ENABLE_EMPIRICAL = True
    ENABLE_RQ1 = True
    ENABLE_RQ2 = True
    ENABLE_RQ3 = True
    base_result_dir = "/root/smart_target/result"
    result_out_dir_name = "out"
    if os.path.exists(os.path.join(base_result_dir, result_out_dir_name)):
        shutil.rmtree(os.path.join(base_result_dir, result_out_dir_name))
    os.mkdir(os.path.join(base_result_dir, result_out_dir_name))
    empirical_result_time = "new/1657702079.5958123:fully:slither:full:open:28800"
    # 单独跑slither的time文件名，如果非特殊情况，应该和empirical_result_time一致
    slither_report_result_time = empirical_result_time
    # RQ1中, slither提供目标制导时的结果, 应该是类似slither:full:open
    rq1_mythril_target_report_result_time = "new/1657775844.6066804:target:slither:full:open:28800"
    # RQ2, 将smartbugs的人工标注的漏洞行号作为目标, 该文件夹只包含target模式的结果, 类似smartbugs:each:open
    rq2_only_target_for_smart_bugs_result_dir = "new/1657849432.6712213:target:smart_bugs:each:open:28800"
    # RQ2里, 单独跑slither的结果, 如无特殊情况，应该和rq2_only_target_for_smart_bugs_result_dir一致
    slither_report_each_result_time = rq2_only_target_for_smart_bugs_result_dir
    # RQ3, 不考虑污点语句的Mythril-target结果, 应该与smart-bugs作为目标的each模式比较, 应该是类似smartbugs:each:close
    rq3_mythril_target_without_state_statement_result_dir = "new/1657858235.1910374:target:smart_bugs:each:close:28800"
    if not ENABLE_EMPIRICAL or not ENABLE_RQ1:  # 实证研究和RQ1在同一个大循环处理里, 不能分割
        exit(-1)
    total_size = 0
    success_size = 0
    reports = dict()
    sol_paths = set()  # 用于记录本次跑出来的实验, 共有多少组合约被跑了, 注意记录的是所有, 不是成功跑出来的
    assert_error_emprial_exam_rq1_list = list()
    # 以下为所谓的原始实验, 包含了slither, Mythril-fully和mythril-target(slihter提供目标)这三个工具的数据
    for root, dirs, files in os.walk(os.path.join(base_result_dir, empirical_result_time)):
        for file in files:
            # 为什么把smart_target.json作为标准? 因为这个文件是最后生成的, 若未生成, 则说明前面的步骤出现了问题, 若生成, 则至少流程走下来了
            if file.endswith("smart_target.json"):
                smart_target_json_path = os.path.join(root, file)
                category_sol_path_contract_name_smart_target_json_list = smart_target_json_path.split(empirical_result_time)[1].split("/")
                assert len(
                    category_sol_path_contract_name_smart_target_json_list) == 4 and \
                       category_sol_path_contract_name_smart_target_json_list[0] == ''
                category = category_sol_path_contract_name_smart_target_json_list[1]
                # print("正在分析:", smart_target_json_path)
                sol_path = smart_target_json_path.replace(base_result_dir, "/root/smartbugs/dataset").replace(empirical_result_time, "").replace("smart_target.json", "").replace("//", "/").split(":")[0] + ".sol"
                assert os.path.exists(sol_path), "源代码文件不存在"

                contract_name = smart_target_json_path.replace(base_result_dir, "/root/smartbugs/dataset").replace(empirical_result_time, "").replace("smart_target.json", "").replace("//", "/").split(":")[1].replace("/", "")
                sol_paths.add(sol_path + ":" + contract_name)
                total_size += 1
                try:
                    mythril_output_fully_json_path = smart_target_json_path.replace("smart_target.json", "mythril_output_fully.json")
                    assert os.path.exists(mythril_output_fully_json_path), "mythril全量执行RESULT不存在"
                    mythril_coverage_fully_txt_path = mythril_output_fully_json_path.replace("mythril_output_fully.json", "coverage_fully.txt")
                    assert os.path.exists(mythril_coverage_fully_txt_path), "mythril全量执行coverage不存在"
                    # mythril-target应该用rq1的结果进行替换, empirical_result_time里不存在RQ1中的target数据
                    mythril_output_target_json_path = smart_target_json_path.replace("smart_target.json", "mythril_output_target.json").replace(empirical_result_time, rq1_mythril_target_report_result_time)
                    assert os.path.exists(mythril_output_target_json_path), "mythril目标执行RESULT不存在"
                    mythril_coverage_target_txt_path = mythril_output_target_json_path.replace("mythril_output_target.json", "coverage_target.txt").replace(empirical_result_time, rq1_mythril_target_report_result_time)
                    assert os.path.exists(mythril_coverage_target_txt_path), "mythril目标执行coverage不存在"
                    # slither的数据通过slither_report_result_time来替换
                    slither_output_json_path = smart_target_json_path.replace("smart_target.json", "slither.json").replace(empirical_result_time, slither_report_result_time)
                    assert os.path.exists(slither_output_json_path), "slither RESULT不存在"

                    # print("======" + sol_path + ":" + contract_name + "======")
                    summary_fully = json.load(open(smart_target_json_path))  # 这是fully的时间
                    # 数据矫正部分 , 因为slither和mythril-target可能用了别的文件夹的数据, 因此他们的summary时间需要矫正
                    summary_slither = json.load(open(smart_target_json_path.replace(empirical_result_time, slither_report_result_time)))
                    summary_fully['slither_used_time'] = summary_slither['slither_used_time']  # 更新纯粹slither的时间
                    summary_mythril_target = json.load(open(smart_target_json_path.replace(empirical_result_time, rq1_mythril_target_report_result_time)))
                    summary_fully['target_gen_used_time'] = summary_mythril_target['target_gen_used_time']  # 从rq1中，更新slither作为目标，生成制导信息的时间
                    summary_fully['mythril_with_target'] = summary_mythril_target['mythril_with_target']  # 从rq1中，更新slither作为目标，mythril运行的时间
                    summary_fully['target_time_out'] = summary_mythril_target['target_time_out']
                    summary_fully['chose_method'] = summary_mythril_target['chose_method']
                    summary_fully['success'] = summary_mythril_target['success']
                    summary_fully['reason'] = summary_mythril_target['reason']
                    # 矫正部分结束
                    s = Summary(summary_fully['slither_used_time'],  # slither检测所用的时间
                                summary_fully['target_gen_used_time'],  # 生成制导信息的时间
                                summary_fully['mythril_with_target'],  # 制导符号执行的时间
                                summary_fully['mythril_fully'],  # 全量符号执行的时间
                                summary_fully['chose_method'],  # 字节码映射源代码所使用的方法
                                summary_fully['success'],  # 是否成功生成制导信息
                                summary_fully['reason'],  # 未生成制导信息的原因
                                summary_fully['target_time_out'],  # 制导符号执行是否超时
                                summary_fully['fully_time_out'])  # 全量符号执行是否超时

                    mythril_target_report = MythrilReport(sol_path,
                                                          contract_name,
                                                          0,  # 静态检测的时间,为0
                                                          0,  # 制导信息的时间,为0
                                                          summary_fully['mythril_with_target'],  # 符号执行的时间, 这里是制导符号执行
                                                          mythril_output_target_json_path,  # 符号执行的报告json所在位置
                                                          "jsonv1",  # 报告类型
                                                          Mode.TARGET,  # 符号执行的模式
                                                          mythril_coverage_target_txt_path,  # 包含覆盖率信息的文件位置
                                                          category)
                    mythril_fully_report = MythrilReport(sol_path,
                                                         contract_name,
                                                         0,
                                                         0,
                                                         summary_fully['mythril_fully'],
                                                         mythril_output_fully_json_path,
                                                         "jsonv1",
                                                         Mode.FULLY,
                                                         mythril_coverage_fully_txt_path, category)
                    slither_report = SlitherReport(sol_path,
                                                   contract_name,
                                                   summary_fully['slither_used_time'],
                                                   0,
                                                   0,
                                                   slither_output_json_path, category)

                    smart_target_report = SmartTargetReport(slither_report, mythril_target_report, mythril_fully_report, s)
                    reports[sol_path + ":" + contract_name] = smart_target_report  # sol_path包含了category，所以无需担心不同类别的重名问题
                    success_size += 1
                    # print("分析成功")
                except AssertionError as assert_error:
                    print(f"断言错误: {assert_error}")
                    assert_error_emprial_exam_rq1_list.append((smart_target_json_path, assert_error))
    assert len(sol_paths) == total_size
    if len(assert_error_emprial_exam_rq1_list) != 0:
        print("实证研究和RQ1大循环过程中出现断言错误, 具体文件名称如下")
        print(assert_error_emprial_exam_rq1_list)
    smart_target_report_path = os.path.join(base_result_dir, empirical_result_time, "smart_target_report.json")
    print(smart_target_report_path)
    print(f"共{total_size}份合约报告文件夹. 成功检测{success_size}.百分比为{success_size / total_size * 100}%")
    smart_target_summary = SmartTargetSummaryReport(reports)  # smart-target-slither数据集合的统计与分析, 这个变量不能被覆盖
    json.dump(smart_target_summary, open(smart_target_report_path, "w"), default=json_outputer, indent=4)
    # 实证研究
    empirical_research_excel_path = os.path.join(base_result_dir, result_out_dir_name, f'{Goal.EMPIRICAL_RESEARCH.name}.xlsx')
    empirical_research_results = gen_results(smart_target_summary, Goal.EMPIRICAL_RESEARCH)
    empirical_research_results = [r.__dict__ for r in empirical_research_results]
    empirical_research_df = pandas.DataFrame(empirical_research_results)
    handle_empirical_research_and_rq1_df_to_excel(empirical_research_df, empirical_research_excel_path)  # 将df处理为excel, 合并表格
    # RQ1
    rq1_excel_path = os.path.join(base_result_dir, result_out_dir_name, f'{Goal.RQ1.name}.xlsx')
    rq1_results = gen_results(smart_target_summary, Goal.RQ1)
    rq1_results = [r.__dict__ for r in rq1_results]
    rq1_df = pandas.DataFrame(rq1_results)
    handle_empirical_research_and_rq1_df_to_excel(rq1_df, rq1_excel_path)
    # RQ1 Without Support, 去掉了Slither制导的影响（数值漏洞）, 去掉了Mythril无法分析的漏洞
    rq1_withour_arithmetic_excel_path = os.path.join(base_result_dir, result_out_dir_name, f'{Goal.RQ1_WITHOUT_NOT_SUPPORT.name}.xlsx')
    rq1_withour_arithmetic_results = gen_results(smart_target_summary, Goal.RQ1_WITHOUT_NOT_SUPPORT)
    rq1_withour_arithmetic_results = [r.__dict__ for r in rq1_withour_arithmetic_results]
    rq1_withour_arithmetic_df = pandas.DataFrame(rq1_withour_arithmetic_results)
    handle_empirical_research_and_rq1_df_to_excel(rq1_withour_arithmetic_df, rq1_withour_arithmetic_excel_path)
    draw_venn_plot_rq1(smart_target_summary)
    if not ENABLE_RQ2:
        exit(-2)
    # RQ2 准备部分, 从跑出来的数据中将smartbugs作为目标的Mythril-target的结果提取出来
    reports_rq2 = dict()
    sol_paths_rq2 = set()
    assert_error_rq2_list = list()
    key_error_rq2_list = list()
    for root, dirs, files in os.walk(os.path.join(base_result_dir, rq2_only_target_for_smart_bugs_result_dir)):
        for file in files:
            if file.endswith("smart_target.json"):
                smart_target_json_path = os.path.join(root, file)
                category_sol_path_contract_name_smart_target_json_list = smart_target_json_path.split(rq2_only_target_for_smart_bugs_result_dir)[1].split("/")
                assert len(
                    category_sol_path_contract_name_smart_target_json_list) == 5 and \
                       category_sol_path_contract_name_smart_target_json_list[0] == ''
                category = category_sol_path_contract_name_smart_target_json_list[1]

                # 第index个人工标注的漏洞, 这个文件夹仅检查这个漏洞
                origin_bug_index = category_sol_path_contract_name_smart_target_json_list[-2]
                # print("正在分析:", smart_target_json_path)
                sol_path = smart_target_json_path.replace(base_result_dir, "/root/smartbugs/dataset").replace(rq2_only_target_for_smart_bugs_result_dir, "").replace("smart_target.json", "").replace("//", "/").split(":")[0] + ".sol"
                assert os.path.exists(sol_path), "源代码文件不存在"
                contract_name = smart_target_json_path.replace(base_result_dir, "/root/smartbugs/dataset").replace(rq2_only_target_for_smart_bugs_result_dir, "").replace("smart_target.json", "").replace("//", "/").split(":")[1].split("/")
                assert origin_bug_index == contract_name[1]
                contract_name = contract_name[0]
                sol_paths_rq2.add(sol_path + ":" + contract_name)
                try:
                    mythril_output_target_json_path = smart_target_json_path.replace("smart_target.json", "mythril_output_target.json")
                    assert os.path.exists(mythril_output_target_json_path), "mythril目标执行RESULT不存在"
                    mythril_coverage_target_txt_path = mythril_output_target_json_path.replace("mythril_output_target.json", "coverage_target.txt")
                    assert os.path.exists(mythril_coverage_target_txt_path), "mythril目标执行coverage不存在"
                    slither_output_json_path = smart_target_json_path.replace("smart_target.json", "slither.json").replace(rq2_only_target_for_smart_bugs_result_dir, slither_report_each_result_time)
                    assert os.path.exists(slither_output_json_path), "slither RESULT不存在"

                    # print("======" + sol_path + ":" + contract_name +":" + origin_bug_index + "======")
                    summary_fully = json.load(open(smart_target_json_path))
                    s = Summary(summary_fully['slither_used_time'],  # slither检测所用的时间, RQ2里这个值没有意义
                                summary_fully['target_gen_used_time'],  # 生成制导信息的时间
                                summary_fully['mythril_with_target'],  # 制导符号执行的时间
                                summary_fully['mythril_fully'],  # 全量符号执行的时间, RQ2里这个值没有意义,肯定是一个特别小的数（应该是0）, 因为fully模式被跳过了
                                summary_fully['chose_method'],  # 字节码映射源代码所使用的方法
                                summary_fully['success'],  # 是否成功生成制导信息
                                summary_fully['reason'],  # 未生成制导信息的原因
                                summary_fully['target_time_out'],  # 制导符号执行是否超时
                                summary_fully['fully_time_out'])  # 全量符号执行是否超时, RQ2里这个值没有意义
                    mythril_target_report = MythrilReport(sol_path,
                                                          contract_name,
                                                          0,  # 静态检测的时间,为0
                                                          0,  # 制导信息的时间,为0
                                                          summary_fully['mythril_with_target'],  # 符号执行的时间, 这里是制导符号执行
                                                          mythril_output_target_json_path,  # 符号执行的报告json所在位置
                                                          "jsonv1",  # 报告类型
                                                          Mode.TARGET,  # 符号执行的模式
                                                          mythril_coverage_target_txt_path,  # 包含覆盖率信息的文件位置
                                                          category,
                                                          BugInfoDetectMode.EACH_TIME,
                                                          smart_target_json_path.replace("smart_target.json", "slither.json")  # 提供slither.json，用于确认哪个漏洞是origin，也就是只检查这个漏洞是否被检测出来了
                                                          )
                    slither_report = SlitherReport(sol_path,
                                                   contract_name,
                                                   summary_fully['slither_used_time'],
                                                   0,
                                                   0,
                                                   slither_output_json_path, category)
                    # 余下的数据, 也就是全量Mythril, 需要从原始数据（实证研究）中寻找，首先，寻找这些数据，加载他们，最后生成report的时候用下面的替换mytrhil-fully的数据
                    mythril_fully_report = smart_target_summary.detail[sol_path + ":" + contract_name].mythtil_fully_report
                    # 为了通过断言检测，只需要对summary里的数据改变就行了, 需要对RQ2中的部分与时间有关的数据进行修改, 将无意义的数据以原始数据替换, 注意！Mythril-target-slither模式的任何数据以RQ2的为准
                    s.slither_time = slither_report.static_time_cost
                    s.mythril_fully_time = mythril_fully_report.dynamic_time_cost
                    # 数据更正, 到此完毕, 不能再进行修改
                    smart_target_report = SmartTargetReport(slither_report, mythril_target_report, mythril_fully_report, s)
                    reports_rq2[sol_path + ":" + contract_name + ":" + origin_bug_index] = smart_target_report
                    # print("RQ2, 分析成功")
                except AssertionError as assert_error:
                    print(f'RQ2, 断言错误: {assert_error}')
                    assert_error_rq2_list.append((smart_target_json_path, assert_error))
                except KeyError as key_error:
                    print(f"RQ2, 出现key error, {key_error}")
                    key_error_rq2_list.append((smart_target_json_path, key_error))
    if len(assert_error_rq2_list) != 0:
        print("断言出现过错误, 具体为如下文件中的断言")
        print(assert_error_rq2_list)
    if len(key_error_rq2_list) != 0:
        print(f"RQ2出现key error的数量不为0, 这些大概率是因为RQ1里面缺少这个的数据")
        print(key_error_rq2_list)
    smart_target_report_path_rq2 = os.path.join(base_result_dir, rq2_only_target_for_smart_bugs_result_dir, "smart_target_report_rq2.json")
    smart_target_summary_rq2 = SmartTargetSummaryReport(reports_rq2, BugInfoDetectMode.EACH_TIME)
    json.dump(smart_target_summary_rq2, open(smart_target_report_path_rq2, "w"), default=json_outputer, indent=4)
    # RQ2 数据分析部分
    if len(sol_paths) != len(sol_paths_rq2):
        print("RQ2, 警告!RQ2和最初的原始实验所检测的合约存在数量不对应的问题!")
        print("原始比RQ2多:", sol_paths - sol_paths_rq2)
        print("RQ2比原始多:", sol_paths_rq2 - sol_paths)

    rq2_excel_path = os.path.join(base_result_dir, result_out_dir_name, f'{Goal.RQ2.name}.xlsx')
    rq2_plot_path = os.path.join(base_result_dir, result_out_dir_name, f'{Goal.RQ2.name}.plot.png')
    rq2_bar_path = os.path.join(base_result_dir, result_out_dir_name, f'{Goal.RQ2.name}.bar.png')
    rq2_results = gen_results(smart_target_summary_rq2, Goal.RQ2)
    rq2_results = [r.__dict__ for r in rq2_results]
    rq2_df = pandas.DataFrame(rq2_results)
    handle_rq2_df_to_excel(rq2_df, rq2_excel_path)
    handle_rq2_df_to_plot(rq2_df, rq2_plot_path, rq2_bar_path, smart_target_summary_rq2, os.path.join(base_result_dir, result_out_dir_name))
    # RQ2 分漏洞类型的表格
    rq2_each_buh_type_excel_path = os.path.join(base_result_dir, result_out_dir_name, f'{Goal.RQ2_EACH_BUG_TYPE.name}.xlsx')
    rq2_each_bug_type_results = gen_results(smart_target_summary_rq2, Goal.RQ2_EACH_BUG_TYPE)
    rq2_each_bug_type_results = [r.__dict__ for r in rq2_each_bug_type_results]
    rq2_each_bug_type_df = pandas.DataFrame(rq2_each_bug_type_results)
    handle_rq2_each_bug_type_df_to_excel(rq2_each_bug_type_df, rq2_each_buh_type_excel_path)
    rq2_reproduction_result = handle_rq2_bug_reproduction(rq2_results)
    rq2_reproduction_result = [r.__dict__ for r in rq2_reproduction_result]
    rq2_reproduction_excel_path = os.path.join(base_result_dir, result_out_dir_name, f'{Goal.RQ2_REPRODUCTION.name}.xlsx')
    rq2_reproduction_df = pandas.DataFrame(rq2_reproduction_result)
    rq2_reproduction_df.to_excel(rq2_reproduction_excel_path)
    if not ENABLE_RQ3:
        exit(-3)
    # RQ3 数据准备部分, 从RQ3跑出来的数据的文件夹里, 将数据读取到Mythril-Fully里, 临时作为存储的介质, 别忘了在输出datafream的时候转换回去
    total_size_rq3 = 0
    success_size_rq3 = 0
    reports_rq3 = dict()
    sol_paths_rq3 = set()
    for root, dirs, files in os.walk(os.path.join(base_result_dir, rq3_mythril_target_without_state_statement_result_dir)):
        for file in files:
            if file.endswith("smart_target.json"):
                smart_target_json_path = os.path.join(root, file)
                category_sol_path_contract_name_smart_target_json_list = smart_target_json_path.split(rq3_mythril_target_without_state_statement_result_dir)[1].split("/")
                assert len(
                    category_sol_path_contract_name_smart_target_json_list) == 5 and \
                       category_sol_path_contract_name_smart_target_json_list[0] == ''
                category = category_sol_path_contract_name_smart_target_json_list[1]
                origin_bug_index = category_sol_path_contract_name_smart_target_json_list[-2]
                # print("正在分析:", smart_target_json_path)
                sol_path = smart_target_json_path.replace(base_result_dir, "/root/smartbugs/dataset").replace(rq3_mythril_target_without_state_statement_result_dir, "").replace("smart_target.json", "").replace("//", "/").split(":")[0] + ".sol"
                assert os.path.exists(sol_path), "源代码文件不存在"
                contract_name = smart_target_json_path.replace(base_result_dir, "/root/smartbugs/dataset").replace(rq3_mythril_target_without_state_statement_result_dir, "").replace("smart_target.json", "").replace("//", "/").split(":")[1].split('/')
                assert origin_bug_index == contract_name[1]
                contract_name = contract_name[0]
                sol_paths_rq3.add(sol_path + ":" + contract_name)
                try:
                    # 虽然下面写的是mythril-target, 但这个之后需要被当作mythril-fully处理, 而mythril-target需要从最上面的RQ2位置拿到
                    mythril_output_target_json_path = smart_target_json_path.replace("smart_target.json", "mythril_output_target.json")
                    assert os.path.exists(mythril_output_target_json_path), "mythril目标执行RESULT不存在"
                    mythril_coverage_target_txt_path = mythril_output_target_json_path.replace("mythril_output_target.json", "coverage_target.txt")
                    assert os.path.exists(mythril_coverage_target_txt_path), "mythril目标执行coverage不存在"
                    # print("======" + sol_path + ":" + contract_name + "======")
                    summary_fully = json.load(open(smart_target_json_path))
                    s = Summary(summary_fully['slither_used_time'],  # slither检测所用的时间, RQ3里这个值没有意义
                                summary_fully['target_gen_used_time'],  # 生成制导信息的时间
                                summary_fully['mythril_with_target'],  # 制导符号执行的时间, 代码到这行暂时先这样, 但实际上这个应该是mythril-target(存在污点分析)的时间
                                # 全量符号执行的时间, RQ3里这个值没有意义,肯定是一个特别小的数, 因为fully模式被跳过了,
                                # 代码到这里, 先暂时这样, 实际上这个值应该被summary['mythril_with_target']代替, 因为我们将mythril-target(无污点分析)得出来的值作为fully
                                summary_fully['mythril_fully'],
                                summary_fully['chose_method'],  # 字节码映射源代码所使用的方法
                                summary_fully['success'],  # 是否成功生成制导信息
                                summary_fully['reason'],  # 未生成制导信息的原因
                                summary_fully['target_time_out'],  # 制导符号执行是否超时
                                summary_fully['fully_time_out'])  # 全量符号执行是否超时, RQ3里这个值没有意义
                    # 注意！虽然是mythril-target(no statement), 但为了实验分析, 将其作为mythril-fully
                    mythril_fully_report = MythrilReport(sol_path,
                                                         contract_name,
                                                         0,  # 静态检测的时间,为0
                                                         0,  # 制导信息的时间,为0
                                                         # 符号执行的时间, 这里是制导符号执行
                                                         # 注意！虽然是mythril-target(no statement), 但为了实验分析, 将其作为mythril-fully
                                                         summary_fully['mythril_with_target'],
                                                         mythril_output_target_json_path,  # 符号执行的报告json所在位置
                                                         "jsonv1",  # 报告类型
                                                         Mode.TARGET,  # 符号执行的模式
                                                         mythril_coverage_target_txt_path,  # 包含覆盖率信息的文件位置
                                                         category,
                                                         BugInfoDetectMode.EACH_TIME,
                                                         smart_target_json_path.replace("smart_target.json", "slither.json"))
                    # 余下的数据, 也就是slither、制导Mythril（RQ2）这两个, 需要从原始数据（RQ2）中寻找
                    # 从smartbugs制导的mythril-target中获得结果, 因为这个RQ3考虑的是smartbugs作为目标来源(也就是原始数据里的mythril-target)
                    mythril_target_report = smart_target_summary_rq2.detail[sol_path + ":" + contract_name + ":" + origin_bug_index].mythril_target_report
                    slither_report = smart_target_summary_rq2.detail[sol_path + ":" + contract_name + ":" + origin_bug_index].slither_report
                    s.slither_time = slither_report.static_time_cost
                    s.mythril_target_time = mythril_target_report.dynamic_time_cost  # target_time应该是开启污点分析的mythril-target时间
                    s.mythril_fully_time = mythril_fully_report.dynamic_time_cost  # fully_time应该是未开启污点分析, 也就是这个循环在做的mythril-fully的时间
                    # 数据更正, 到此完毕, 不能再进行修改了, 此时, fully为未开启污点分析, target为开启污点分析, slither没有任何含义
                    smart_target_report = SmartTargetReport(slither_report, mythril_target_report, mythril_fully_report, s)
                    reports_rq3[sol_path + ":" + contract_name + ":" + origin_bug_index] = smart_target_report
                    success_size_rq3 += 1
                except AssertionError as assert_error:
                    print(f"RQ3, 断言错误:{assert_error}")
    smart_target_report_path_rq3 = os.path.join(base_result_dir, rq3_mythril_target_without_state_statement_result_dir, "smart_target_report_rq3.json")
    smart_target_summary_rq3 = SmartTargetSummaryReport(reports_rq3, BugInfoDetectMode.EACH_TIME)
    json.dump(smart_target_summary_rq3, open(smart_target_report_path_rq3, "w"), default=json_outputer, indent=4)
    # RQ3 数据分析与Excel生成
    if len(sol_paths) != len(sol_paths_rq3):
        print("RQ3, 警告!RQ3和最初的原始实验所检测的合约存在数量不对应的问题!")
        print("原始比RQ3多:", sol_paths - sol_paths_rq3)
        print("RQ3比原始多:", sol_paths_rq3 - sol_paths)
    rq3_excel_path = os.path.join(base_result_dir, result_out_dir_name, f'{Goal.RQ3.name}.xlsx')
    rq3_plot_path = os.path.join(base_result_dir, result_out_dir_name, f'{Goal.RQ3.name}.png')
    rq3_results = gen_results(smart_target_summary_rq3, Goal.RQ3)
    rq3_results = [r.__dict__ for r in rq3_results]
    rq3_df = pandas.DataFrame(rq3_results)
    handle_rq2_df_to_excel(rq3_df, rq3_excel_path, True)
    handle_rq3_result_to_plot(rq3_results, os.path.join(base_result_dir, result_out_dir_name), rq1_results)
    rq3_state_write_result = gen_results(smart_target_summary_rq3, Goal.RQ3_WITH_STATE_WRITE_INFO)
    rq3_state_write_result = [r.__dict__ for r in rq3_state_write_result]
    rq3_state_write_excel_path = os.path.join(base_result_dir, result_out_dir_name, f'{Goal.RQ3_WITH_STATE_WRITE_INFO.name}.xlsx')
    rq3_state_write_df = pandas.DataFrame(rq3_state_write_result)
    rq3_state_write_df.to_excel(rq3_state_write_excel_path, index=False)
    verify_result_path = os.path.join(base_result_dir, result_out_dir_name)
    # verify_data_complete_and_threat_analysis("/root/smartbugs/dataset",
    #                                          os.path.join(base_result_dir, empirical_result_time),
    #                                          os.path.join(base_result_dir, rq1_mythril_target_report_result_time),
    #                                          os.path.join(base_result_dir, rq2_only_target_for_smart_bugs_result_dir),
    #                                          os.path.join(base_result_dir, rq3_mythril_target_without_state_statement_result_dir),
    #                                          verify_result_path)
